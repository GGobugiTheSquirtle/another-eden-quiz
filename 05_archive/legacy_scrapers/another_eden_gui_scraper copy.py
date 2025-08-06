import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, unquote, parse_qs, urlparse
import os
import time
import re
import csv
import mimetypes
import configparser
import sys
import shutil
from eden_data_preprocess_gui import make_roulette_csv_from_excel

# --- 기본 설정 ---
BASE_URL = "https://anothereden.wiki"
TARGET_URL = "https://anothereden.wiki/w/Characters"
CONFIG_FILE = "scraper_config.ini"
NAME_MAPPING_FILE = "Matching_names.csv"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_BASE_DIR = SCRIPT_DIR
IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")
EXCEL_FILENAME_BASE = "another_eden_characters_detailed"

# --- 설정 파일 및 폴더 처리 (이전과 동일) ---
def load_config():
    global OUTPUT_BASE_DIR, IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV
    config = configparser.ConfigParser()
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    if os.path.exists(config_path):
        config.read(config_path)
        OUTPUT_BASE_DIR = config.get('Paths', 'OutputDirectory', fallback=SCRIPT_DIR)
    else:
        OUTPUT_BASE_DIR = SCRIPT_DIR
    
    IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")

def save_config():
    global OUTPUT_BASE_DIR
    config = configparser.ConfigParser()
    config['Paths'] = {'OutputDirectory': OUTPUT_BASE_DIR}
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    with open(config_path, 'w') as configfile:
        config.write(configfile)

def setup_directories():
    if not os.path.exists(IMAGE_DIR):
        os.makedirs(IMAGE_DIR)
    icons_path = os.path.join(IMAGE_DIR, "icons")
    elements_path = os.path.join(IMAGE_DIR, "elements_equipment")
    if not os.path.exists(icons_path):
        os.makedirs(icons_path)
    if not os.path.exists(elements_path):
        os.makedirs(elements_path)

def get_unique_filename(filepath):
    if not os.path.exists(filepath):
        return filepath
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base} ({counter}){ext}"
        if not os.path.exists(new_filepath):
            return new_filepath
        counter += 1

def download_image(image_url, subfolder=""):
    if not image_url: return None
    full_image_url = urljoin(BASE_URL, image_url)
    try:
        parsed_url = urlparse(full_image_url)
        query_params = parse_qs(parsed_url.query)
        image_name_from_f = query_params.get('f', [None])[0]
        
        if image_name_from_f:
            image_name = os.path.basename(unquote(image_name_from_f))
        else:
            image_name = os.path.basename(unquote(parsed_url.path.split('?')[0]))

        if not image_name or image_name.lower() in ["thumb.php", "index.php"]:
            temp_name = unquote(full_image_url.split('/')[-1].split('?')[0])
            image_name = (temp_name[:50] + ".png") if temp_name else "unknown_image.png"

        base_name, ext = os.path.splitext(image_name)
        if not ext or len(ext) > 5:
            try:
                head_resp = requests.head(full_image_url, timeout=3, allow_redirects=True)
                head_resp.raise_for_status()
                content_type = head_resp.headers.get('Content-Type')
                if content_type:
                    guessed_ext = mimetypes.guess_extension(content_type.split(';')[0])
                    image_name = base_name + (guessed_ext if guessed_ext and guessed_ext != '.jpe' else ('.jpg' if guessed_ext == '.jpe' else '.png'))
                else: image_name = base_name + ".png" 
            except: image_name = base_name + ".png"
        
        image_name = re.sub(r'[<>:"/\\|?*]', '_', image_name)
        image_name = image_name[:200]

        save_path_dir = os.path.join(IMAGE_DIR, subfolder)
        save_path = os.path.join(save_path_dir, image_name)
        
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            return save_path

        img_response = requests.get(full_image_url, stream=True, timeout=10)
        img_response.raise_for_status()
        with open(save_path, 'wb') as f:
            for chunk in img_response.iter_content(8192):
                f.write(chunk)
        time.sleep(0.05) 
        return save_path
    except requests.exceptions.RequestException as e:
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue: error_log_queue.put(f"Download Error (Network) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    except Exception as e:
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue: error_log_queue.put(f"Download Error (Other) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    return None

def load_name_mapping(log_queue_ref):
    mapping_dict = {}
    mapping_file_path = os.path.join(SCRIPT_DIR, NAME_MAPPING_FILE)
    if not os.path.exists(mapping_file_path):
        log_queue_ref.put(f"Warning: Name mapping file not found at '{mapping_file_path}'. Using original names.", "error")
        return mapping_dict
    try:
        with open(mapping_file_path, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.reader(infile)
            next(reader)
            for rows in reader:
                if len(rows) >= 2:
                    eng_name = rows[0].strip()
                    kor_name = rows[1].strip()
                    if eng_name:
                        mapping_dict[eng_name] = kor_name
        log_queue_ref.put(f"Successfully loaded {len(mapping_dict)} name mappings.", "success")
    except Exception as e:
        log_queue_ref.put(f"Error loading name mapping file: {e}", "error")
    return mapping_dict

# --- 스크레이핑 핵심 로직 ---
def scraping_logic(log_queue_ref, progress_queue_ref, generate_structure_sheet_only_mode, selected_output_dir, final_excel_full_path):
    threading.current_thread().log_queue_ref = log_queue_ref
    threading.current_thread().progress_queue_ref = progress_queue_ref
    
    global IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV 
    IMAGE_DIR = os.path.join(selected_output_dir, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(selected_output_dir, "structure_analysis.csv")
    
    setup_directories()
    
    log_queue_ref.put(f"Output directory set to: {selected_output_dir}")
    name_map = load_name_mapping(log_queue_ref)

    log_queue_ref.put(f"Fetching page: {TARGET_URL}")
    headers_ua = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    try:
        response = requests.get(TARGET_URL, headers=headers_ua, timeout=15)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        log_queue_ref.put(f"Error fetching page: {e}")
        progress_queue_ref.put({'done': True, 'error': True})
        return

    log_queue_ref.put("Page fetched. Parsing HTML...")
    soup = BeautifulSoup(response.content, 'html.parser')
    
    char_table = soup.find('table', class_='chara-table') or soup.find('table', class_='wikitable')
    if not char_table:
        log_queue_ref.put("Character table not found. Please verify table class name.")
        progress_queue_ref.put({'done': True, 'error': True})
        return

    log_queue_ref.put("Character table found. Parsing rows...")
    
    all_character_data_for_final_excel = []
    rows = char_table.find_all('tr')
    total_rows_to_process = len(rows) - 1 
    log_queue_ref.put(f"Found {total_rows_to_process} potential character rows.")
    if total_rows_to_process <= 0 :
        log_queue_ref.put("No character rows found to process.")
        progress_queue_ref.put({'done': True})
        return
        
    progress_queue_ref.put({'max': total_rows_to_process, 'value': 0})

    for i, row in enumerate(rows[1:]): 
        current_progress = i + 1
        progress_queue_ref.put({'value': current_progress}) 
        cells = row.find_all('td')
        if len(cells) < 4: continue

        if generate_structure_sheet_only_mode:
            if current_progress % 50 == 0: log_queue_ref.put(f"Structure Analysis: Processed row {current_progress}/{total_rows_to_process}")
            continue
        
        try:
            icon_cell = cells[0]
            icon_img_tag = icon_cell.find('img')
            icon_src = icon_img_tag['src'] if icon_img_tag and icon_img_tag.get('src') else None
            icon_local_path = download_image(icon_src, "icons") if icon_src else None
            
            icon_filename = ""
            if icon_src:
                parsed_url = urlparse(icon_src)
                query_params = parse_qs(parsed_url.query)
                image_name_from_f = query_params.get('f', [None])[0]
                if image_name_from_f: icon_filename = os.path.basename(unquote(image_name_from_f)).replace(' ', '_')
                else: icon_filename = os.path.basename(unquote(parsed_url.path.split('?')[0])).replace(' ', '_')

            name_rarity_cell = cells[1]
            name_tag = name_rarity_cell.find('a')
            
            original_name = name_tag.text.strip() if name_tag else ""
            rarity = ""
            lines_in_cell = [line.strip() for line in name_rarity_cell.get_text(separator='\n').splitlines() if line.strip()]
            if not original_name and lines_in_cell: original_name = lines_in_cell[0]
            for line_text in reversed(lines_in_cell):
                if "★" in line_text:
                    rarity = line_text
                    break
            if not rarity:
                rarity_match = re.search(r'\d(?:~\d)?★(?:\s*\S+)?', name_rarity_cell.get_text(separator=" ").strip())
                if rarity_match: rarity = rarity_match.group(0).strip()
            
            final_name = name_map.get(original_name, original_name)
            
            element_equipment_cell = cells[2]
            ee_icon_tags = element_equipment_cell.find_all('img')
            element_equipment_icon_paths = []
            element_equipment_icon_alts = []
            for img_tag in ee_icon_tags:
                ee_src = img_tag.get('src')
                ee_alt = img_tag.get('alt', "") 
                if ee_src:
                    local_path = download_image(ee_src, "elements_equipment")
                    if local_path:
                        element_equipment_icon_paths.append(local_path)
                        element_equipment_icon_alts.append(ee_alt)

            release_date = cells[3].text.strip()

            if final_name or icon_local_path:
                all_character_data_for_final_excel.append({
                    "icon_path": icon_local_path, "icon_filename": icon_filename, "name": final_name,
                    "rarity": rarity, "element_equipment_paths": element_equipment_icon_paths,
                    "element_equipment_alts": element_equipment_icon_alts, "release_date": release_date
                })
                log_name = final_name if final_name != original_name else original_name
                if current_progress % 20 == 0 : log_queue_ref.put(f"Row {current_progress}: Parsed & Downloaded for '{log_name}'")

        except Exception as e:
            log_queue_ref.put(f"Row {current_progress}: Error parsing for final Excel: {e}")
            continue
    
    if not generate_structure_sheet_only_mode:
        if not all_character_data_for_final_excel:
            log_queue_ref.put("No data to save to the final Excel sheet.")
            progress_queue_ref.put({'done': True})
            return

        # ★★★★★ 변경점 1: 이미지 정리를 위한 데이터 정렬 및 폴더 생성 ★★★★★
        log_queue_ref.put("Organizing character images...")
        organized_base_dir = os.path.join(selected_output_dir, "캐릭터 정리")
        by_date_dir = os.path.join(organized_base_dir, "1. 출시일 순")
        by_name_dir = os.path.join(organized_base_dir, "2. 가나다 순")
        os.makedirs(by_date_dir, exist_ok=True)
        os.makedirs(by_name_dir, exist_ok=True)

        # 출시일 순으로 정렬 (날짜 형식 'YYYY/MM/DD'를 기준으로)
        # 날짜가 없는 경우 맨 뒤로 보내기 위해 '9999/99/99' 사용
        sorted_by_date = sorted(all_character_data_for_final_excel, key=lambda x: x.get('release_date', '9999/99/99'))
        
        # 가나다 순으로 정렬
        sorted_by_name = sorted(all_character_data_for_final_excel, key=lambda x: x.get('name', ''))

        def copy_and_rename_image(char_data, index, target_dir):
            if char_data["icon_path"] and os.path.exists(char_data["icon_path"]):
                try:
                    rarity_transformed = char_data['rarity'].strip().replace(" ", "").replace("★", "성").replace("SA", "성각") or "정보없음"
                    date_transformed = char_data['release_date'].replace('/', '-') if char_data['release_date'] else "날짜없음"
                    base_filename = f"{char_data['name']}_{rarity_transformed}_{date_transformed}"
                    sanitized_filename_base = re.sub(r'[<>:"/\\|?*]', '_', base_filename)
                    _, extension = os.path.splitext(char_data["icon_path"])
                    if not extension: extension = '.png'
                    
                    # 번호(넘버링) 추가
                    final_organized_filename = f"{index+1:03d}_{sanitized_filename_base}{extension}"
                    
                    destination_path = os.path.join(target_dir, final_organized_filename)
                    shutil.copy2(char_data["icon_path"], destination_path)
                except Exception as e:
                    log_queue_ref.put(f"Failed to copy organized image for '{char_data['name']}': {e}", "error")

        # 정렬된 목록을 기반으로 파일 복사 및 이름 변경 실행
        for idx, data in enumerate(sorted_by_date):
            copy_and_rename_image(data, idx, by_date_dir)
        
        for idx, data in enumerate(sorted_by_name):
            copy_and_rename_image(data, idx, by_name_dir)
            
        log_queue_ref.put("Finished organizing images.", "success")
        
        # --- 엑셀 저장 로직 (이전과 동일) ---
        HEADER_MAP = {
            "Icon": "아이콘", "Icon Filename": "아이콘 파일명", "Name": "이름", "Rarity": "희귀도",
            "Elem/Equip Icon": "속성/장비", "Elem/Equip Alt": "속성/장비 정보", "Release Date": "출시일"
        }
        log_queue_ref.put(f"Saving final data to Excel: {final_excel_full_path} ({len(all_character_data_for_final_excel)} characters)")
        wb = Workbook()
        ws = wb.active
        ws.title = "Characters"
        max_ee_icons = max((len(d["element_equipment_paths"]) for d in all_character_data_for_final_excel), default=0)
        
        headers_excel = [HEADER_MAP["Icon"], HEADER_MAP["Icon Filename"], HEADER_MAP["Name"], HEADER_MAP["Rarity"]]
        for i in range(max_ee_icons):
            headers_excel.extend([f"{HEADER_MAP['Elem/Equip Icon']} {i+1}", f"{HEADER_MAP['Elem/Equip Alt']} {i+1}"])
        headers_excel.append(HEADER_MAP["Release Date"])
        ws.append(headers_excel)

        ws.column_dimensions[get_column_letter(1)].width = 12 
        ws.column_dimensions[get_column_letter(2)].width = 35 
        ws.column_dimensions[get_column_letter(3)].width = 25 
        ws.column_dimensions[get_column_letter(4)].width = 15
        col_offset = 5
        for i in range(max_ee_icons):
            ws.column_dimensions[get_column_letter(col_offset + i*2)].width = 12 
            ws.column_dimensions[get_column_letter(col_offset + i*2 + 1)].width = 25 
        ws.column_dimensions[get_column_letter(col_offset + max_ee_icons*2)].width = 15
        
        for row_idx_excel, char_data in enumerate(all_character_data_for_final_excel, start=2):
            ws.row_dimensions[row_idx_excel].height = 60 
            current_col = 1 
            if char_data["icon_path"] and os.path.exists(char_data["icon_path"]):
                try:
                    img = OpenpyxlImage(char_data["icon_path"])
                    img.height = 75 
                    img.width = 75  
                    ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
                except Exception as e:
                    log_queue_ref.put(f"Excel Write Error (Icon) for {char_data.get('name', 'N/A')}: {e}")
                    ws.cell(row=row_idx_excel, column=current_col, value="ImgErr")
            current_col += 1
            
            ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("icon_filename", ""))
            current_col += 1
            ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("name", "N/A"))
            current_col += 1
            ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("rarity", "N/A"))
            current_col += 1

            for i in range(max_ee_icons):
                if i < len(char_data["element_equipment_paths"]):
                    icon_path = char_data["element_equipment_paths"][i]
                    icon_alt = char_data["element_equipment_alts"][i]
                    if icon_path and os.path.exists(icon_path):
                        try:
                            img = OpenpyxlImage(icon_path)
                            img.height = 30 
                            img.width = 30
                            ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
                        except Exception as e:
                            log_queue_ref.put(f"Excel Write Error (E/E Icon) for {char_data.get('name', 'N/A')}: {e}")
                            ws.cell(row=row_idx_excel, column=current_col, value="ImgErr")
                    ws.cell(row=row_idx_excel, column=current_col + 1, value=icon_alt)
                else: 
                    ws.cell(row=row_idx_excel, column=current_col, value="")
                    ws.cell(row=row_idx_excel, column=current_col + 1, value="")
                current_col += 2 
            
            ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("release_date", "N/A"))
            log_name = char_data.get('name')
            if (row_idx_excel-1) % 50 == 0 : log_queue_ref.put(f"Excel: Wrote data for '{log_name}' (Processed {row_idx_excel-1} characters)")

        try:
            wb.save(final_excel_full_path)
            log_queue_ref.put(f"Final data successfully saved to {final_excel_full_path}")
            try:
                csv_generated_path = make_roulette_csv_from_excel(final_excel_full_path)
                log_queue_ref.put(f"Roulette CSV generated: {csv_generated_path}")
            except Exception as e_pre:
                log_queue_ref.put(f"Error generating roulette CSV: {e_pre}")
        except Exception as e:
            log_queue_ref.put(f"Error saving final Excel file: {e}")
            progress_queue_ref.put({'done': True, 'error': True, 'error_message': f"Excel 저장 오류:\n{e}\n\n파일이 다른 프로그램에서 열려있는지 확인하세요."})
            return

    progress_queue_ref.put({'done': True})


class ScraperApp:
    # --- GUI 클래스 로직 (이전과 동일, 변경 없음) ---
    def __init__(self, root_window):
        self.root = root_window
        root_window.title("어나더에덴 캐릭터 스크레이퍼")
        root_window.geometry("800x600")

        load_config()
        self.output_dir_var = tk.StringVar(value=OUTPUT_BASE_DIR)

        style = ttk.Style()
        try:
            if os.name == 'nt': style.theme_use('vista')
            else: style.theme_use('clam') 
        except tk.TclError: style.theme_use('clam')

        main_frame = ttk.Frame(root_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        path_frame = ttk.LabelFrame(main_frame, text="출력 설정", padding="10")
        path_frame.pack(fill=tk.X, pady=(0,10))

        ttk.Label(path_frame, text="저장 폴더:").grid(row=0, column=0, padx=(0,5), pady=5, sticky="w")
        self.output_dir_entry = ttk.Entry(path_frame, textvariable=self.output_dir_var, width=60)
        self.output_dir_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.btn_browse_dir = ttk.Button(path_frame, text="폴더 찾아보기", command=self.browse_output_directory)
        self.btn_browse_dir.grid(row=0, column=2, padx=(5,0), pady=5)
        path_frame.columnconfigure(1, weight=1)

        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0,10))

        self.btn_structure = ttk.Button(control_frame, text="1. 구조 분석 시트 생성", command=lambda: self.start_scraping_thread(structure_only=True))
        self.btn_structure.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        self.btn_full_report = ttk.Button(control_frame, text="2. 최종 보고서 생성 (이미지 포함)", command=lambda: self.start_scraping_thread(structure_only=False))
        self.btn_full_report.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        
        self.btn_open_folder = ttk.Button(control_frame, text="출력 폴더 열기", command=self.open_output_folder)
        self.btn_open_folder.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        log_label_frame = ttk.LabelFrame(main_frame, text="진행 상황 및 로그", padding="10")
        log_label_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(log_label_frame, wrap=tk.WORD, state=tk.DISABLED, height=15, font=("Malgun Gothic", 9) if os.name == 'nt' else ("TkDefaultFont", 10))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.tag_configure("error", foreground="red")
        self.log_text.tag_configure("info", foreground="blue")
        self.log_text.tag_configure("success", foreground="green")

        self.progress_bar = ttk.Progressbar(log_label_frame, orient="horizontal", length=300, mode="determinate")
        self.progress_bar.pack(pady=10, fill=tk.X, padx=5)

        self.log_queue = queue.Queue()
        self.progress_queue = queue.Queue()
        self.root.after(100, self.process_queues)
        
        root_window.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        save_config()
        self.root.destroy()

    def browse_output_directory(self):
        global OUTPUT_BASE_DIR
        selected_dir = filedialog.askdirectory(initialdir=self.output_dir_var.get())
        if selected_dir:
            self.output_dir_var.set(selected_dir)
            OUTPUT_BASE_DIR = selected_dir
            self.log_message(f"출력 폴더가 다음으로 설정되었습니다: {selected_dir}", "info")
            save_config()

    def open_output_folder(self):
        folder_to_open = self.output_dir_var.get()
        if not os.path.isdir(folder_to_open):
            self.log_message(f"폴더를 찾을 수 없습니다: {folder_to_open}", "error")
            messagebox.showerror("오류", f"폴더를 찾을 수 없습니다:\n{folder_to_open}")
            return
        try:
            if os.name == 'nt': os.startfile(folder_to_open)
            elif sys.platform == "darwin": os.system(f'open "{folder_to_open}"')
            else: os.system(f'xdg-open "{folder_to_open}"')
            self.log_message(f"출력 폴더가 열렸습니다: {folder_to_open}", "info")
        except Exception as e:
            self.log_message(f"폴더 열기 오류: {e}", "error")
            messagebox.showerror("폴더 열기 오류", f"출력 폴더를 여는 중 오류 발생:\n{e}")

    def log_message(self, message, tag=None):
        self.log_text.config(state=tk.NORMAL)
        timestamp = time.strftime("%H:%M:%S", time.localtime())
        if tag: self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", tag)
        else: self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END) 

    def start_scraping_thread(self, structure_only):
        current_output_dir = self.output_dir_var.get()
        if not os.path.isdir(current_output_dir):
            messagebox.showerror("오류", f"유효한 출력 폴더를 선택해주세요.\n현재 경로: {current_output_dir}")
            return

        final_excel_path_for_thread = ""
        
        if not structure_only:
            target_path = os.path.join(current_output_dir, f"{EXCEL_FILENAME_BASE}.xlsx")
            
            if os.path.exists(target_path):
                choice = messagebox.askyesnocancel("파일 확인", f"'{os.path.basename(target_path)}' 파일이 이미 존재합니다.\n\n▶ 예: 기존 파일을 덮어씁니다.\n▶ 아니오: 새 이름으로 저장합니다.\n▶ 취소: 작업을 중단합니다.")
                if choice is True:
                    final_excel_path_for_thread = target_path
                    self.log_message(f"기존 파일을 덮어쓸 예정입니다: {target_path}", "info")
                elif choice is False:
                    final_excel_path_for_thread = get_unique_filename(target_path)
                    self.log_message(f"새 이름으로 저장될 예정입니다: {final_excel_path_for_thread}", "info")
                else:
                    self.log_message("작업이 사용자에 의해 취소되었습니다.", "info")
                    return
            else:
                final_excel_path_for_thread = target_path
        
        self.btn_structure.config(state=tk.DISABLED)
        self.btn_full_report.config(state=tk.DISABLED)
        self.progress_bar["value"] = 0
        self.log_text.config(state=tk.NORMAL); self.log_text.delete('1.0', tk.END); self.log_text.config(state=tk.DISABLED)
        self.log_message("Scraping process started...", "info")

        save_config()

        self.thread = threading.Thread(target=scraping_logic, args=(self.log_queue, self.progress_queue, structure_only, current_output_dir, final_excel_path_for_thread), daemon=True)
        self.thread.start()

    def process_queues(self):
        try:
            while True: 
                message = self.log_queue.get_nowait()
                tag_to_use = None
                if "Error" in message or "오류" in message: tag_to_use = "error"
                elif "saved" in message or "완료" in message or "successfully" in message: tag_to_use = "success"
                elif "Fetching" in message or "Parsing" in message or "set to" in message: tag_to_use = "info"
                self.log_message(message, tag_to_use)
        except queue.Empty: pass

        try:
            while True: 
                progress_data = self.progress_queue.get_nowait()
                if 'max' in progress_data: self.progress_bar["maximum"] = progress_data['max']
                if 'value' in progress_data: self.progress_bar["value"] = progress_data['value']
                if 'done' in progress_data and progress_data['done']:
                    self.btn_structure.config(state=tk.NORMAL)
                    self.btn_full_report.config(state=tk.NORMAL)
                    if self.progress_bar["value"] < self.progress_bar["maximum"]: self.progress_bar["value"] = self.progress_bar["maximum"] 
                    
                    if 'error' in progress_data and progress_data['error']:
                        msg = progress_data.get('error_message', "작업 중 오류가 발생했습니다. 로그를 확인하세요.")
                        self.log_message(f"Process finished with errors: {msg}", "error")
                        messagebox.showerror("오류", msg)
                    else:
                        self.log_message("Process finished successfully.", "success")
                        messagebox.showinfo("완료", "작업이 성공적으로 완료되었습니다!")
        except queue.Empty: pass
        self.root.after(100, self.process_queues)

if __name__ == "__main__":
    root = tk.Tk()
    app = ScraperApp(root)
    root.mainloop()