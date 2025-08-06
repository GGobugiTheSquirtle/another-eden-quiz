import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, unquote, parse_qs, urlparse
import os
import time
import re
import csv
import mimetypes
import configparser
import sys
# from eden_data_preprocess_gui import make_roulette_csv_from_excel # 이 부분은 이제 스트림릿에서 처리하므로 주석 처리하거나 삭제 가능

# --- 기본 설정 ---
BASE_URL = "https://anothereden.wiki"
TARGET_URL = "https://anothereden.wiki/w/Characters"
CONFIG_FILE = "scraper_config.ini"

# 전역 변수 초기화
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_BASE_DIR = SCRIPT_DIR
IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")
EXCEL_FILENAME_BASE = "another_eden_characters_detailed"

# --- 설정 파일 처리 ---
def load_config():
    global OUTPUT_BASE_DIR, IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV
    config = configparser.ConfigParser()
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    if os.path.exists(config_path):
        config.read(config_path)
        OUTPUT_BASE_DIR = config.get('Paths', 'OutputDirectory', fallback=SCRIPT_DIR)
    else:
        OUTPUT_BASE_DIR = SCRIPT_DIR
    
    IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")

def save_config():
    global OUTPUT_BASE_DIR
    config = configparser.ConfigParser()
    config['Paths'] = {'OutputDirectory': OUTPUT_BASE_DIR}
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    with open(config_path, 'w') as configfile:
        config.write(configfile)

# --- 이미지 및 폴더 처리 함수 (기존과 동일) ---
def setup_directories():
    if not os.path.exists(IMAGE_DIR):
        os.makedirs(IMAGE_DIR)
    icons_path = os.path.join(IMAGE_DIR, "icons")
    elements_path = os.path.join(IMAGE_DIR, "elements_equipment")
    if not os.path.exists(icons_path):
        os.makedirs(icons_path)
    if not os.path.exists(elements_path):
        os.makedirs(elements_path)

def get_unique_filename(filepath):
    if not os.path.exists(filepath):
        return filepath
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base} ({counter}){ext}"
        if not os.path.exists(new_filepath):
            return new_filepath
        counter += 1

def download_image(image_url, subfolder=""):
    if not image_url:
        return None
    full_image_url = urljoin(BASE_URL, image_url)
    try:
        parsed_url = urlparse(full_image_url)
        query_params = parse_qs(parsed_url.query)
        image_name_from_f = query_params.get('f', [None])[0]
        
        if image_name_from_f:
            image_name = os.path.basename(unquote(image_name_from_f))
        else:
            image_name = os.path.basename(unquote(parsed_url.path.split('?')[0]))

        if not image_name or image_name.lower() in ["thumb.php", "index.php"]:
            temp_name = unquote(full_image_url.split('/')[-1].split('?')[0])
            image_name = (temp_name[:50] + ".png") if temp_name else "unknown_image.png"

        base_name, ext = os.path.splitext(image_name)
        if not ext or len(ext) > 5:
            try:
                head_resp = requests.head(full_image_url, timeout=3, allow_redirects=True)
                head_resp.raise_for_status()
                content_type = head_resp.headers.get('Content-Type')
                if content_type:
                    guessed_ext = mimetypes.guess_extension(content_type.split(';')[0])
                    image_name = base_name + (guessed_ext if guessed_ext and guessed_ext != '.jpe' else '.jpg')
                else: image_name = base_name + ".png" 
            except: image_name = base_name + ".png"
        
        image_name = re.sub(r'[<>:"/\\|?*]', '_', image_name)[:200]

        save_path_dir = os.path.join(IMAGE_DIR, subfolder)
        save_path = os.path.join(save_path_dir, image_name)
        
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            return save_path

        img_response = requests.get(full_image_url, stream=True, timeout=10)
        img_response.raise_for_status()
        with open(save_path, 'wb') as f:
            for chunk in img_response.iter_content(8192):
                f.write(chunk)
        time.sleep(0.05) 
        return save_path
    except requests.exceptions.RequestException as e:
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue: error_log_queue.put(f"Download Error (Network) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    except Exception as e:
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue: error_log_queue.put(f"Download Error (Other) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    return None

# --- 스크레이핑 핵심 로직 ---
def scraping_logic(log_queue_ref, progress_queue_ref, generate_structure_sheet_only_mode, selected_output_dir):
    threading.current_thread().log_queue_ref = log_queue_ref
    threading.current_thread().progress_queue_ref = progress_queue_ref
    
    global IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV 
    IMAGE_DIR = os.path.join(selected_output_dir, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(selected_output_dir, "structure_analysis.csv")
    
    setup_directories()
    log_queue_ref.put(f"Output directory set to: {selected_output_dir}")
    log_queue_ref.put(f"Fetching page: {TARGET_URL}")
    
    headers_ua = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    try:
        response = requests.get(TARGET_URL, headers=headers_ua, timeout=15)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        log_queue_ref.put(f"Error fetching page: {e}")
        progress_queue_ref.put({'done': True, 'error': True})
        return

    log_queue_ref.put("Page fetched. Parsing HTML...")
    soup = BeautifulSoup(response.content, 'html.parser')
    
    char_table = soup.find('table', class_='chara-table') or soup.find('table', class_='wikitable')
    if not char_table:
        log_queue_ref.put("Character table not found.")
        progress_queue_ref.put({'done': True, 'error': True})
        return

    log_queue_ref.put("Character table found. Parsing rows...")
    
    all_character_data_for_final_excel = []
    raw_data_for_structure_analysis = []
    
    rows = char_table.find_all('tr')
    total_rows_to_process = len(rows) - 1
    log_queue_ref.put(f"Found {total_rows_to_process} potential character rows.")
    if total_rows_to_process <= 0:
        log_queue_ref.put("No character rows found.")
        progress_queue_ref.put({'done': True})
        return
        
    progress_queue_ref.put({'max': total_rows_to_process, 'value': 0})

    for i, row in enumerate(rows[1:]): 
        current_progress = i + 1
        progress_queue_ref.put({'value': current_progress}) 

        cells = row.find_all('td')
        if len(cells) < 4: continue

        # (구조 분석용 데이터 수집은 기존과 동일)
        if generate_structure_sheet_only_mode:
            # 구조 분석 모드일 때는 상세 페이지 로딩 없이 빠르게 진행
            # ... (기존 구조 분석 코드) ...
            continue

        try:
            # 1. 메인 테이블 정보 추출
            icon_cell = cells[0]
            icon_img_tag = icon_cell.find('img')
            icon_src = icon_img_tag['src'] if icon_img_tag else None
            icon_local_path = download_image(icon_src, "icons") if icon_src else None
            
            icon_filename = ""
            if icon_src:
                parsed = urlparse(icon_src)
                fname = parse_qs(parsed.query).get('f', [None])[0] or os.path.basename(unquote(parsed.path))
                icon_filename = fname.replace(' ', '_')

            name_rarity_cell = cells[1]
            name_tag = name_rarity_cell.find('a')
            name = name_tag.text.strip() if name_tag else ""
            
            rarity_text = name_rarity_cell.get_text(separator=" ").strip()
            rarity_match = re.search(r'\d(?:~\d)?★(?:\s*\S+)?', rarity_text)
            rarity = rarity_match.group(0).strip() if rarity_match else ""

            element_equipment_cell = cells[2]
            ee_icon_tags = element_equipment_cell.find_all('img')
            element_equipment_icon_paths, element_equipment_icon_alts = [], []
            for img_tag in ee_icon_tags:
                ee_src, ee_alt = img_tag.get('src'), img_tag.get('alt', "")
                if ee_src:
                    local_path = download_image(ee_src, "elements_equipment")
                    if local_path:
                        element_equipment_icon_paths.append(local_path)
                        element_equipment_icon_alts.append(ee_alt)
            
            release_date = cells[3].text.strip()

            ## ================== [추가] Personalities 스크레이핑 시작 ==================
            personalities = []
            detail_href = name_tag.get('href') if name_tag else None
            if detail_href:
                detail_url = urljoin(BASE_URL, detail_href)
                try:
                    detail_resp = requests.get(detail_url, headers=headers_ua, timeout=10)
                    detail_resp.raise_for_status()
                    detail_soup = BeautifulSoup(detail_resp.content, 'html.parser')
                    
                    # 'Personalities' id를 가진 span 태그 찾기
                    headline = detail_soup.find('span', id='Personalities')
                    if headline:
                        # span의 부모(h2 또는 h3)를 찾고, 그 다음 형제인 ul 태그를 찾음
                        parent_heading = headline.find_parent(['h2', 'h3'])
                        ul_tag = parent_heading.find_next_sibling('ul') if parent_heading else None
                        if ul_tag:
                            # li 태그의 텍스트만 추출 (이미지 alt 텍스트나 숫자 등 포함될 수 있음)
                            # 예: "Mask  ... 1" -> "Mask 1" 으로 추출됨
                            personalities = [li.get_text(strip=True) for li in ul_tag.find_all('li')]
                except Exception as e:
                    log_queue_ref.put(f"Error fetching personalities for '{name}': {e}")
            ## ================== [추가] Personalities 스크레이핑 종료 ==================

            if name or icon_local_path:
                all_character_data_for_final_excel.append({
                    "icon_path": icon_local_path,
                    "icon_filename": icon_filename,
                    "name": name,
                    "rarity": rarity,
                    "personalities": personalities, # [추가] 추출한 퍼스널리티 저장
                    "element_equipment_paths": element_equipment_icon_paths,
                    "element_equipment_alts": element_equipment_icon_alts,
                    "release_date": release_date
                })
                if current_progress % 10 == 0:
                    log_queue_ref.put(f"Row {current_progress}: Parsed & Downloaded for '{name}'")
        except Exception as e:
            log_queue_ref.put(f"Row {current_progress}: Error parsing for final Excel: {e}")
            continue

    if generate_structure_sheet_only_mode:
        log_queue_ref.put("Structure analysis mode finished.")
        progress_queue_ref.put({'done': True})
        return

    if not all_character_data_for_final_excel:
        log_queue_ref.put("No data to save to Excel.")
        progress_queue_ref.put({'done': True})
        return

    # --- 엑셀 파일 생성 ---
    final_excel_path_base = os.path.join(selected_output_dir, EXCEL_FILENAME_BASE)
    final_excel_full_path = get_unique_filename(f"{final_excel_path_base}.xlsx")
    log_queue_ref.put(f"Saving data to Excel: {final_excel_full_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Characters"
    
    max_ee_icons = max([len(d["element_equipment_paths"]) for d in all_character_data_for_final_excel] or [0])

    ## ================== [수정] 엑셀 헤더에 Personalities 추가 ==================
    headers_excel = ["Icon", "Icon Filename", "Name", "Rarity", "Personalities"]
    for i in range(max_ee_icons):
        headers_excel.extend([f"Elem/Equip {i+1} Icon", f"Elem/Equip {i+1} Alt"])
    headers_excel.append("Release Date")
    ws.append(headers_excel)

    # 컬럼 너비 설정
    ws.column_dimensions[get_column_letter(1)].width = 12  # Icon
    ws.column_dimensions[get_column_letter(2)].width = 35  # Icon Filename
    ws.column_dimensions[get_column_letter(3)].width = 30  # Name
    ws.column_dimensions[get_column_letter(4)].width = 15  # Rarity
    ws.column_dimensions[get_column_letter(5)].width = 40  # [추가] Personalities
    ee_start_col = 6
    for i in range(max_ee_icons):
        ws.column_dimensions[get_column_letter(ee_start_col + i*2)].width = 12 
        ws.column_dimensions[get_column_letter(ee_start_col + i*2 + 1)].width = 25 
    ws.column_dimensions[get_column_letter(ee_start_col + max_ee_icons*2)].width = 15

    for row_idx_excel, char_data in enumerate(all_character_data_for_final_excel, start=2):
        ws.row_dimensions[row_idx_excel].height = 60 
        
        # Icon
        current_col = 1
        if char_data["icon_path"] and os.path.exists(char_data["icon_path"]):
            try:
                img = OpenpyxlImage(char_data["icon_path"])
                img.height, img.width = 75, 75
                ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
            except Exception: pass
        current_col += 1
        
        # Icon Filename, Name, Rarity
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("icon_filename", "")); current_col += 1
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("name", "")); current_col += 1
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("rarity", "")); current_col += 1
        
        ## ================== [수정] 엑셀 셀에 Personalities 데이터 쓰기 ==================
        personalities_text = ", ".join(char_data.get("personalities", []))
        ws.cell(row=row_idx_excel, column=current_col, value=personalities_text)
        current_col += 1
        
        # Elem/Equip
        for i in range(max_ee_icons):
            if i < len(char_data["element_equipment_paths"]):
                icon_path = char_data["element_equipment_paths"][i]
                icon_alt = char_data["element_equipment_alts"][i]
                if icon_path and os.path.exists(icon_path):
                    try:
                        img = OpenpyxlImage(icon_path)
                        img.height, img.width = 30, 30
                        ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
                    except Exception: pass
                ws.cell(row=row_idx_excel, column=current_col + 1, value=icon_alt)
            current_col += 2
        
        # Release Date
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("release_date", ""))
        
    try:
        wb.save(final_excel_full_path)
        log_queue_ref.put(f"Excel file saved successfully: {final_excel_full_path}")
    except Exception as e:
        log_queue_ref.put(f"Error saving Excel file: {e}")
        progress_queue_ref.put({'done': True, 'error': True, 'error_message': f"Excel 저장 오류:\n{e}"})
        return
    
    progress_queue_ref.put({'done': True})

# --- Tkinter GUI Application (기존과 거의 동일, eden_data_preprocess_gui 임포트 부분만 주석 처리) ---
class ScraperApp:
    def __init__(self, root_window):
        self.root = root_window
        root_window.title("어나더에덴 캐릭터 스크레이퍼")
        root_window.geometry("800x600")

        load_config()
        self.output_dir_var = tk.StringVar(value=OUTPUT_BASE_DIR)

        style = ttk.Style()
        try:
            if os.name == 'nt': style.theme_use('vista')
            else: style.theme_use('clam') 
        except tk.TclError: style.theme_use('clam')

        main_frame = ttk.Frame(root_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        path_frame = ttk.LabelFrame(main_frame, text="출력 설정", padding="10")
        path_frame.pack(fill=tk.X, pady=(0,10))
        
        ttk.Label(path_frame, text="저장 폴더:").grid(row=0, column=0, padx=(0,5), pady=5, sticky="w")
        self.output_dir_entry = ttk.Entry(path_frame, textvariable=self.output_dir_var, width=60)
        self.output_dir_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.btn_browse_dir = ttk.Button(path_frame, text="폴더 찾아보기", command=self.browse_output_directory)
        self.btn_browse_dir.grid(row=0, column=2, padx=(5,0), pady=5)
        path_frame.columnconfigure(1, weight=1)

        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0,10))

        # self.btn_structure = ttk.Button(control_frame, text="1. 구조 분석 시트 생성", command=lambda: self.start_scraping_thread(True))
        # self.btn_structure.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        self.btn_full_report = ttk.Button(control_frame, text="캐릭터 데이터 생성 (이미지 및 퍼스널리티 포함)", command=lambda: self.start_scraping_thread(False))
        self.btn_full_report.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        self.btn_open_folder = ttk.Button(control_frame, text="출력 폴더 열기", command=self.open_output_folder)
        self.btn_open_folder.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        log_label_frame = ttk.LabelFrame(main_frame, text="진행 상황 및 로그", padding="10")
        log_label_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(log_label_frame, wrap=tk.WORD, state=tk.DISABLED, height=15, font=("Malgun Gothic", 9) if os.name == 'nt' else ("TkDefaultFont", 10))
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.tag_configure("error", foreground="red")
        self.log_text.tag_configure("info", foreground="blue")
        self.log_text.tag_configure("success", foreground="green")

        self.progress_bar = ttk.Progressbar(log_label_frame, orient="horizontal", length=300, mode="determinate")
        self.progress_bar.pack(pady=10, fill=tk.X, padx=5)

        self.log_queue = queue.Queue()
        self.progress_queue = queue.Queue()
        self.root.after(100, self.process_queues)
        
        root_window.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        save_config()
        self.root.destroy()

    def browse_output_directory(self):
        global OUTPUT_BASE_DIR
        selected_dir = filedialog.askdirectory(initialdir=self.output_dir_var.get())
        if selected_dir:
            self.output_dir_var.set(selected_dir)
            OUTPUT_BASE_DIR = selected_dir
            self.log_message(f"출력 폴더가 다음으로 설정되었습니다: {selected_dir}", "info")
            save_config()

    def open_output_folder(self):
        folder_to_open = self.output_dir_var.get()
        if not os.path.isdir(folder_to_open):
            messagebox.showerror("오류", f"폴더를 찾을 수 없습니다:\n{folder_to_open}")
            return
        try:
            if os.name == 'nt': os.startfile(folder_to_open)
            elif sys.platform == "darwin": os.system(f'open "{folder_to_open}"')
            else: os.system(f'xdg-open "{folder_to_open}"')
        except Exception as e:
            messagebox.showerror("폴더 열기 오류", f"폴더를 여는 중 오류 발생:\n{e}")

    def log_message(self, message, tag=None):
        self.log_text.config(state=tk.NORMAL)
        timestamp = time.strftime("%H:%M:%S", time.localtime())
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", tag)
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END) 

    def start_scraping_thread(self, structure_only_mode):
        current_output_dir = self.output_dir_var.get()
        if not os.path.isdir(current_output_dir):
            messagebox.showerror("오류", f"유효한 출력 폴더를 선택해주세요.")
            return

        # self.btn_structure.config(state=tk.DISABLED)
        self.btn_full_report.config(state=tk.DISABLED)
        self.progress_bar["value"] = 0
        self.log_text.config(state=tk.NORMAL); self.log_text.delete('1.0', tk.END); self.log_text.config(state=tk.DISABLED)
        self.log_message("Scraping process started...", "info")

        save_config()

        self.thread = threading.Thread(target=scraping_logic, args=(self.log_queue, self.progress_queue, structure_only_mode, current_output_dir), daemon=True)
        self.thread.start()

    def process_queues(self):
        try:
            while True: 
                message = self.log_queue.get_nowait()
                tag = "error" if "Error" in message or "오류" in message else "success" if "saved" in message or "완료" in message else "info"
                self.log_message(message, tag)
        except queue.Empty: pass

        try:
            while True: 
                progress_data = self.progress_queue.get_nowait()
                if 'max' in progress_data: self.progress_bar["maximum"] = progress_data['max']
                if 'value' in progress_data: self.progress_bar["value"] = progress_data['value']
                if progress_data.get('done'):
                    # self.btn_structure.config(state=tk.NORMAL)
                    self.btn_full_report.config(state=tk.NORMAL)
                    if self.progress_bar["value"] < self.progress_bar["maximum"]:
                        self.progress_bar["value"] = self.progress_bar["maximum"] 
                    
                    if progress_data.get('error'):
                        msg = progress_data.get('error_message', "작업 중 오류 발생.")
                        messagebox.showerror("오류", msg)
                    else:
                        messagebox.showinfo("완료", "작업이 성공적으로 완료되었습니다!")
        except queue.Empty: pass
        self.root.after(100, self.process_queues)

if __name__ == "__main__":
    root = tk.Tk()
    app = ScraperApp(root)
    root.mainloop()