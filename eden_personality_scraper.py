"""
🔧 Another Eden Personality 데이터 스크레이퍼
기존 캐릭터 데이터에 Personalities 정보를 추가하는 개선된 스크레이퍼
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, unquote, parse_qs, urlparse
import os
import time
import re
# import csv  # unused in current implementation
import mimetypes
import configparser
import sys
import pandas as pd

# --- 기본 설정 ---
BASE_URL = "https://anothereden.wiki"
TARGET_URL = "https://anothereden.wiki/w/Characters"
PERSONALITY_URL = "https://anothereden.wiki/w/Characters/Personality"
CONFIG_FILE = "scraper_config.ini"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_BASE_DIR = SCRIPT_DIR
IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")
EXCEL_FILENAME_BASE = "another_eden_characters_detailed"

# --- 설정 파일 처리 ---
def load_config():
    global OUTPUT_BASE_DIR, IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV
    config = configparser.ConfigParser()
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    if os.path.exists(config_path):
        config.read(config_path)
        OUTPUT_BASE_DIR = config.get('Paths', 'OutputDirectory', fallback=SCRIPT_DIR)
    else:
        OUTPUT_BASE_DIR = SCRIPT_DIR
    
    IMAGE_DIR = os.path.join(OUTPUT_BASE_DIR, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(OUTPUT_BASE_DIR, "structure_analysis.csv")

def save_config():
    global OUTPUT_BASE_DIR
    config = configparser.ConfigParser()
    config['Paths'] = {'OutputDirectory': OUTPUT_BASE_DIR}
    config_path = os.path.join(SCRIPT_DIR, CONFIG_FILE)
    with open(config_path, 'w') as configfile:
        config.write(configfile)

def setup_directories():
    if not os.path.exists(IMAGE_DIR):
        os.makedirs(IMAGE_DIR)
    icons_path = os.path.join(IMAGE_DIR, "icons")
    elements_path = os.path.join(IMAGE_DIR, "elements_equipment")
    if not os.path.exists(icons_path):
        os.makedirs(icons_path)
    if not os.path.exists(elements_path):
        os.makedirs(elements_path)

def get_unique_filename(filepath):
    if not os.path.exists(filepath):
        return filepath
    base, ext = os.path.splitext(filepath)
    counter = 1
    while True:
        new_filepath = f"{base} ({counter}){ext}"
        if not os.path.exists(new_filepath):
            return new_filepath
        counter += 1

def download_image(image_url, subfolder=""):
    if not image_url:
        return None
    full_image_url = urljoin(BASE_URL, image_url)
    try:
        parsed_url = urlparse(full_image_url)
        query_params = parse_qs(parsed_url.query)
        image_name_from_f = query_params.get('f', [None])[0]
        
        if image_name_from_f:
            image_name = os.path.basename(unquote(image_name_from_f))
        else:
            image_name = os.path.basename(unquote(parsed_url.path.split('?')[0]))

        if not image_name or image_name.lower() in ["thumb.php", "index.php"]:
            temp_name = unquote(full_image_url.split('/')[-1].split('?')[0])
            image_name = (temp_name[:50] + ".png") if temp_name else "unknown_image.png"

        base_name, ext = os.path.splitext(image_name)
        if not ext or len(ext) > 5:
            try:
                head_resp = requests.head(full_image_url, timeout=3, allow_redirects=True)
                head_resp.raise_for_status()
                content_type = head_resp.headers.get('Content-Type')
                if content_type:
                    guessed_ext = mimetypes.guess_extension(content_type.split(';')[0])
                    image_name = base_name + (guessed_ext if guessed_ext and guessed_ext != '.jpe' else '.jpg')
                else: 
                    image_name = base_name + ".png" 
            except Exception: 
                image_name = base_name + ".png"
        
        image_name = re.sub(r'[<>:"/\\|?*]', '_', image_name)[:200]

        save_path_dir = os.path.join(IMAGE_DIR, subfolder)
        save_path = os.path.join(save_path_dir, image_name)
        
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            return save_path

        img_response = requests.get(full_image_url, stream=True, timeout=10)
        img_response.raise_for_status()
        with open(save_path, 'wb') as f:
            for chunk in img_response.iter_content(8192):
                f.write(chunk)
        time.sleep(0.05) 
        return save_path
    except requests.exceptions.RequestException as e:
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue: 
            error_log_queue.put(f"Download Error (Network) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    except Exception as e:
        error_log_queue = getattr(threading.current_thread(), 'log_queue_ref', None)
        if error_log_queue: 
            error_log_queue.put(f"Download Error (Other) for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    return None

# --- 퍼스널리티 매칭 데이터 로드 ---
def load_personality_matching():
    """퍼스널리티 매칭 CSV 파일을 로드합니다."""
    personality_mapping = {}
    try:
        personality_csv_path = os.path.join(SCRIPT_DIR, "personality_matching.csv")
        if os.path.exists(personality_csv_path):
            df = pd.read_csv(personality_csv_path)
            for _, row in df.iterrows():
                if 'English' in row and 'Korean' in row:
                    personality_mapping[row['English']] = row['Korean']
    except Exception as e:
        print(f"Error loading personality matching: {e}")
    return personality_mapping

def scrape_all_personalities_from_page(headers_ua, log_queue_ref):
    """
    Characters/Personality 페이지에서 전체 퍼스널리티 목록을 가져옵니다.
    
    Returns:
        dict: {character_name: [personalities]} 형태의 딕셔너리
    """
    personality_mapping = load_personality_matching()
    name_mapping = load_character_name_matching()
    character_personalities = {}
    
    try:
        log_queue_ref.put(f"Fetching personality data from: {PERSONALITY_URL}")
        response = requests.get(PERSONALITY_URL, headers=headers_ua, timeout=15)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # wikitable 클래스를 가진 테이블들 찾기
        tables = soup.find_all('table', class_='wikitable')
        log_queue_ref.put(f"Found {len(tables)} personality tables")
        
        for table_idx, table in enumerate(tables):
            rows = table.find_all('tr')[1:]  # 헤더 제외
            
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    # 첫 번째 셀: 퍼스널리티 이름
                    personality_cell = cells[0]
                    personality_eng = personality_cell.get_text(strip=True)
                    
                    # 한국어 퍼스널리티명으로 변환
                    personality_kor = personality_mapping.get(personality_eng, personality_eng)
                    
                    # 두 번째 셀: 캐릭터 목록
                    characters_cell = cells[1]
                    
                    # 캐릭터 링크들 추출
                    character_links = characters_cell.find_all('a', href=True)
                    
                    for link in character_links:
                        href = link.get('href', '')
                        # 캐릭터 페이지인지 확인 (/w/로 시작하고 특정 패턴 제외)
                        if (href.startswith('/w/') and 
                            'Character' not in href and 
                            'Personality' not in href and
                            'Special:' not in href and
                            'Category:' not in href):
                            
                            char_name = link.get_text(strip=True)
                            if char_name and len(char_name) > 1:
                                # 스타일 접미사를 고려한 캐릭터명 정리 (AS/ES 등 제거)
                                # 딕셔너리 키는 기본 영어명으로 통일 (매칭 시 일관성을 위해)
                                base_char_name = char_name
                                style_patterns = [r'\s+(AS)$', r'\s+(ES)$', r'\s+(NS)$', 
                                                r'\s+(Another\s*Style)$', r'\s+(Extra\s*Style)$']
                                for pattern in style_patterns:
                                    match = re.search(pattern, char_name, re.IGNORECASE)
                                    if match:
                                        base_char_name = char_name[:match.start()]
                                        break
                                
                                # 캐릭터에 퍼스널리티 추가 (기본명을 키로 사용)
                                if base_char_name not in character_personalities:
                                    character_personalities[base_char_name] = []
                                if personality_kor not in character_personalities[base_char_name]:
                                    character_personalities[base_char_name].append(personality_kor)
        
        log_queue_ref.put(f"Found personality data for {len(character_personalities)} characters")
        
    except Exception as e:
        log_queue_ref.put(f"Error scraping personality page: {e}")
    
    return character_personalities

def scrape_personalities(detail_url, headers_ua, log_queue_ref):
    """
    캐릭터 상세 페이지에서 Personalities 정보를 추출합니다.
    
    Returns:
        list: 추출된 퍼스널리티 리스트
    """
    personality_mapping = load_personality_matching()
    personalities = []
    try:
        detail_resp = requests.get(detail_url, headers=headers_ua, timeout=10)
        detail_resp.raise_for_status()
        detail_soup = BeautifulSoup(detail_resp.content, 'html.parser')
        
        # 여러 방법으로 퍼스널리티 찾기
        
        # 방법 1: 'Personalities' id를 가진 span 태그 찾기
        headline = detail_soup.find('span', id='Personalities')
        if headline:
            parent_heading = headline.find_parent(['h2', 'h3'])
            if parent_heading:
                # ul 태그 찾기
                ul_tag = parent_heading.find_next_sibling('ul')
                if ul_tag:
                    for li in ul_tag.find_all('li'):
                        link = li.find('a')
                        if link:
                            personality_eng = link.get_text(strip=True)
                            personality_kor = personality_mapping.get(personality_eng, personality_eng)
                            if personality_kor and personality_kor not in personalities:
                                personalities.append(personality_kor)
        
        # 방법 2: 'Personality' 텍스트가 포함된 헤딩 찾기
        if not personalities:
            for heading in detail_soup.find_all(['h2', 'h3', 'h4']):
                if 'personalit' in heading.get_text().lower():
                    next_element = heading.find_next_sibling()
                    while next_element:
                        if next_element.name == 'ul':
                            for li in next_element.find_all('li'):
                                link = li.find('a')
                                if link:
                                    personality_eng = link.get_text(strip=True)
                                    personality_kor = personality_mapping.get(personality_eng, personality_eng)
                                    if personality_kor and personality_kor not in personalities:
                                        personalities.append(personality_kor)
                            break
                        elif next_element.name in ['h2', 'h3', 'h4']:
                            break
                        next_element = next_element.find_next_sibling()
        
        # 방법 3: 페이지에서 'Personality:' 텍스트 직접 찾기
        if not personalities:
            for element in detail_soup.find_all(text=re.compile(r'personalit', re.IGNORECASE)):
                parent = element.parent
                if parent:
                    # 주변에서 링크 찾기
                    nearby_links = parent.find_parent().find_all('a', href=True) if parent.find_parent() else []
                    for link in nearby_links:
                        href = link.get('href', '')
                        if '/w/Personalities/' in href or 'personality' in href.lower():
                            personality_eng = link.get_text(strip=True)
                            personality_kor = personality_mapping.get(personality_eng, personality_eng)
                            if personality_kor and personality_kor not in personalities:
                                personalities.append(personality_kor)
        
        if personalities:
            log_queue_ref.put(f"Found personalities: {', '.join(personalities)}")
        else:
            # 디버그 정보: 페이지에 실제로 Personalities 관련 내용이 있는지 확인
            personality_mentions = detail_soup.find_all(text=re.compile(r'personalit', re.IGNORECASE))
            if personality_mentions:
                log_queue_ref.put(f"Page mentions 'personality' {len(personality_mentions)} times but no structured data found")
            else:
                log_queue_ref.put("No personalities found on this page")
            
    except Exception as e:
        log_queue_ref.put(f"Error scraping personalities: {e}")
    
    return personalities

# --- 캐릭터명 매칭 데이터 로드 ---
def load_character_name_matching():
    """Matching_names.csv에서 캐릭터명 매칭 데이터를 로드합니다."""
    name_mapping = {}
    try:
        matching_csv_path = os.path.join(SCRIPT_DIR, "Matching_names.csv")
        if os.path.exists(matching_csv_path):
            df = pd.read_csv(matching_csv_path, encoding='utf-8-sig')
            for _, row in df.iterrows():
                if len(row) >= 2:
                    eng_name = row.iloc[0]  # 첫 번째 컬럼 (영어명)
                    kor_name = row.iloc[1]  # 두 번째 컬럼 (한국어명)
                    if pd.notna(eng_name) and pd.notna(kor_name):
                        name_mapping[eng_name] = kor_name
    except Exception as e:
        print(f"Error loading character name matching: {e}")
    return name_mapping

def convert_character_name_with_style(english_name, name_mapping):
    """
    캐릭터명을 한국어로 변환하되, AS/ES/NS 등의 스타일 접미사를 고려합니다.
    예: "Shion ES" → "시온 ES", "Aldo AS" → "알도 AS"
    
    Args:
        english_name (str): 영어 캐릭터명 (스타일 포함 가능)
        name_mapping (dict): 영어-한국어 캐릭터명 매핑 딕셔너리
    
    Returns:
        str: 한국어로 변환된 캐릭터명
    """
    if not english_name:
        return english_name
    
    # 스타일 접미사 패턴 (AS, ES, NS, Another Style 등)
    style_patterns = [
        r'\s+(AS)$',           # " AS"
        r'\s+(ES)$',           # " ES" 
        r'\s+(NS)$',           # " NS"
        r'\s+(Another\s*Style)$',  # " Another Style"
        r'\s+(Extra\s*Style)$',    # " Extra Style"
        r'\s+(Manifestation)$',    # " Manifestation"
        r'\s+(Alter)$',            # " Alter"
    ]
    
    # 스타일 접미사 찾기
    base_name = english_name
    style_suffix = ""
    
    for pattern in style_patterns:
        match = re.search(pattern, english_name, re.IGNORECASE)
        if match:
            style_suffix = " " + match.group(1)
            base_name = english_name[:match.start()]
            break
    
    # 기본 이름을 한국어로 변환
    korean_base_name = name_mapping.get(base_name, base_name)
    
    # 한국어 기본 이름 + 스타일 접미사
    return korean_base_name + style_suffix

def scraping_logic_with_personalities(log_queue_ref, progress_queue_ref, selected_output_dir):
    """개선된 스크레이핑 로직 - 전체 퍼스널리티 페이지와 개별 캐릭터 페이지 병행"""
    threading.current_thread().log_queue_ref = log_queue_ref
    threading.current_thread().progress_queue_ref = progress_queue_ref
    
    global IMAGE_DIR, STRUCTURE_ANALYSIS_FILENAME_CSV 
    IMAGE_DIR = os.path.join(selected_output_dir, "character_art")
    STRUCTURE_ANALYSIS_FILENAME_CSV = os.path.join(selected_output_dir, "structure_analysis.csv")
    
    setup_directories()
    log_queue_ref.put(f"Output directory set to: {selected_output_dir}")
    log_queue_ref.put("Loading character name and personality matching data...")
    
    # 캐릭터명 매칭 데이터 로드
    name_mapping = load_character_name_matching()
    log_queue_ref.put(f"Loaded {len(name_mapping)} character name mappings")
    
    headers_ua = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    
    # 1. 전체 퍼스널리티 페이지에서 데이터 가져오기
    log_queue_ref.put("Step 1: Fetching personality data from comprehensive page...")
    all_personalities_data = scrape_all_personalities_from_page(headers_ua, log_queue_ref)
    
    # 2. 메인 캐릭터 페이지 스크래핑
    log_queue_ref.put(f"Step 2: Fetching main character data from: {TARGET_URL}")
    try:
        response = requests.get(TARGET_URL, headers=headers_ua, timeout=15)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        log_queue_ref.put(f"Error fetching page: {e}")
        progress_queue_ref.put({'done': True, 'error': True})
        return

    log_queue_ref.put("Page fetched. Parsing HTML...")
    soup = BeautifulSoup(response.content, 'html.parser')
    
    char_table = soup.find('table', class_='chara-table') or soup.find('table', class_='wikitable')
    if not char_table:
        log_queue_ref.put("Character table not found.")
        progress_queue_ref.put({'done': True, 'error': True})
        return

    log_queue_ref.put("Character table found. Parsing rows...")
    
    all_character_data_for_final_excel = []
    rows = char_table.find_all('tr')
    total_rows_to_process = len(rows) - 1
    
    log_queue_ref.put(f"Found {total_rows_to_process} potential character rows.")
    if total_rows_to_process <= 0:
        log_queue_ref.put("No character rows found.")
        progress_queue_ref.put({'done': True})
        return
        
    progress_queue_ref.put({'max': total_rows_to_process, 'value': 0})

    for i, row in enumerate(rows[1:]): 
        current_progress = i + 1
        progress_queue_ref.put({'value': current_progress}) 

        cells = row.find_all('td')
        if len(cells) < 4: 
            continue

        try:
            # 1. 메인 테이블 정보 추출
            icon_cell = cells[0]
            icon_img_tag = icon_cell.find('img')
            icon_src = icon_img_tag['src'] if icon_img_tag else None
            icon_local_path = download_image(icon_src, "icons") if icon_src else None
            
            icon_filename = ""
            if icon_src:
                parsed = urlparse(icon_src)
                fname = parse_qs(parsed.query).get('f', [None])[0] or os.path.basename(unquote(parsed.path))
                icon_filename = fname.replace(' ', '_')

            name_rarity_cell = cells[1]
            name_tag = name_rarity_cell.find('a')
            name = name_tag.text.strip() if name_tag else ""
            
            # 캐릭터명 매칭 (영어 -> 한국어, 스타일 접미사 고려)
            korean_name = convert_character_name_with_style(name, name_mapping)
            
            rarity_text = name_rarity_cell.get_text(separator=" ").strip()
            rarity_match = re.search(r'\d(?:~\d)?★(?:\s*\S+)?', rarity_text)
            rarity = rarity_match.group(0).strip() if rarity_match else ""

            element_equipment_cell = cells[2]
            ee_icon_tags = element_equipment_cell.find_all('img')
            element_equipment_icon_paths, element_equipment_icon_alts = [], []
            for img_tag in ee_icon_tags:
                ee_src, ee_alt = img_tag.get('src'), img_tag.get('alt', "")
                if ee_src:
                    local_path = download_image(ee_src, "elements_equipment")
                    if local_path:
                        element_equipment_icon_paths.append(local_path)
                        element_equipment_icon_alts.append(ee_alt)
            
            release_date = cells[3].text.strip()

            # 2. Personalities 데이터 가져오기
            personalities = []
            
            # 기본 캐릭터명 추출 (AS/ES 등 스타일 접미사 제거)
            base_name = name
            style_patterns = [r'\s+(AS)$', r'\s+(ES)$', r'\s+(NS)$', 
                            r'\s+(Another\s*Style)$', r'\s+(Extra\s*Style)$']
            for pattern in style_patterns:
                match = re.search(pattern, name, re.IGNORECASE)
                if match:
                    base_name = name[:match.start()]
                    break
            
            # 우선 전체 퍼스널리티 페이지에서 찾기 (기본명으로 검색)
            if base_name in all_personalities_data:
                personalities = all_personalities_data[base_name]
                log_queue_ref.put(f"Found personalities from main page for '{name}' (base: '{base_name}'): {', '.join(personalities)}")
            else:
                # 개별 캐릭터 페이지에서 스크래핑 (백업)
                detail_href = name_tag.get('href') if name_tag else None
                if detail_href:
                    detail_url = urljoin(BASE_URL, detail_href)
                    personalities = scrape_personalities(detail_url, headers_ua, log_queue_ref)
                    time.sleep(0.5)  # 서버 부담 줄이기 위한 지연

            if name or icon_local_path:
                all_character_data_for_final_excel.append({
                    "icon_path": icon_local_path,
                    "icon_filename": icon_filename,
                    "name": name,
                    "korean_name": korean_name,  # 한국어 캐릭터명 추가
                    "rarity": rarity,
                    "personalities": personalities,
                    "element_equipment_paths": element_equipment_icon_paths,
                    "element_equipment_alts": element_equipment_icon_alts,
                    "release_date": release_date
                })
                if current_progress % 10 == 0:
                    log_queue_ref.put(f"Row {current_progress}: Parsed & Downloaded for '{korean_name}' ({name}) - Personalities: {len(personalities)}")
        except Exception as e:
            log_queue_ref.put(f"Row {current_progress}: Error parsing: {e}")
            continue

    if not all_character_data_for_final_excel:
        log_queue_ref.put("No data to save to Excel.")
        progress_queue_ref.put({'done': True})
        return

    # --- 엑셀 파일 생성 (Personalities 컬럼 포함) ---
    final_excel_path_base = os.path.join(selected_output_dir, EXCEL_FILENAME_BASE)
    final_excel_full_path = get_unique_filename(f"{final_excel_path_base}.xlsx")
    log_queue_ref.put(f"Saving data to Excel: {final_excel_full_path}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Characters"
    
    max_ee_icons = max([len(d["element_equipment_paths"]) for d in all_character_data_for_final_excel] or [0])

    # 헤더에 Korean Name과 Personalities 추가
    headers_excel = ["Icon", "Icon Filename", "Name", "Korean Name", "Rarity", "Personalities"]
    for i in range(max_ee_icons):
        headers_excel.extend([f"Elem/Equip {i+1} Icon", f"Elem/Equip {i+1} Alt"])
    headers_excel.append("Release Date")
    ws.append(headers_excel)

    # 컬럼 너비 설정
    ws.column_dimensions[get_column_letter(1)].width = 12  # Icon
    ws.column_dimensions[get_column_letter(2)].width = 35  # Icon Filename
    ws.column_dimensions[get_column_letter(3)].width = 30  # Name
    ws.column_dimensions[get_column_letter(4)].width = 30  # Korean Name
    ws.column_dimensions[get_column_letter(5)].width = 15  # Rarity
    ws.column_dimensions[get_column_letter(6)].width = 50  # Personalities (넓게)
    
    ee_start_col = 6
    for i in range(max_ee_icons):
        ws.column_dimensions[get_column_letter(ee_start_col + i*2)].width = 12 
        ws.column_dimensions[get_column_letter(ee_start_col + i*2 + 1)].width = 25 
    ws.column_dimensions[get_column_letter(ee_start_col + max_ee_icons*2)].width = 15

    for row_idx_excel, char_data in enumerate(all_character_data_for_final_excel, start=2):
        ws.row_dimensions[row_idx_excel].height = 60 
        
        # Icon
        current_col = 1
        if char_data["icon_path"] and os.path.exists(char_data["icon_path"]):
            try:
                img = OpenpyxlImage(char_data["icon_path"])
                img.height, img.width = 75, 75
                ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
            except Exception: 
                pass
        current_col += 1
        
        # Icon Filename, Name, Korean Name, Rarity, Personalities
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("icon_filename", ""))
        current_col += 1
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("name", ""))
        current_col += 1
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("korean_name", ""))
        current_col += 1
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("rarity", ""))
        current_col += 1
        
        # Personalities 데이터 추가
        personalities_text = ", ".join(char_data.get("personalities", []))
        ws.cell(row=row_idx_excel, column=current_col, value=personalities_text)
        current_col += 1
        
        # Elem/Equip
        for i in range(max_ee_icons):
            if i < len(char_data["element_equipment_paths"]):
                icon_path = char_data["element_equipment_paths"][i]
                icon_alt = char_data["element_equipment_alts"][i]
                if icon_path and os.path.exists(icon_path):
                    try:
                        img = OpenpyxlImage(icon_path)
                        img.height, img.width = 30, 30
                        ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
                    except Exception: 
                        pass
                ws.cell(row=row_idx_excel, column=current_col + 1, value=icon_alt)
            current_col += 2
        
        # Release Date
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("release_date", ""))
        
    try:
        wb.save(final_excel_full_path)
        log_queue_ref.put(f"Excel file saved successfully: {final_excel_full_path}")
        
        # CSV 생성도 함께 수행
        generate_csv_with_personalities(final_excel_full_path, selected_output_dir, log_queue_ref)
        
    except Exception as e:
        log_queue_ref.put(f"Error saving Excel file: {e}")
        progress_queue_ref.put({'done': True, 'error': True, 'error_message': f"Excel 저장 오류:\n{e}"})
        return
    
    progress_queue_ref.put({'done': True})

def generate_csv_with_personalities(excel_path, output_dir, log_queue_ref):
    """엑셀에서 CSV 생성 (Personalities 포함)"""
    try:
        df = pd.read_excel(excel_path, engine='openpyxl')
        
        # 이름 매칭 파일 로드
        name_map = {}
        try:
            name_map_df = pd.read_csv("Matching_names.csv")
            name_map = dict(zip(name_map_df["캐릭터명 (입력)"], name_map_df["캐릭터명 (매칭)"]))
        except FileNotFoundError:
            log_queue_ref.put("Warning: Matching_names.csv not found. Using English names.")
        
        # CSV 데이터 생성
        result = []
        for _, row in df.iterrows():
            name = str(row.get('Name', '')).strip()
            if not name:
                continue
                
            # 한글 이름 변환
            korean_name = name_map.get(name, name)
            
            result.append({
                "캐릭터명": korean_name,
                "희귀도": str(row.get('Rarity', '')).strip(),
                "캐릭터아이콘경로": "",  # 이미지 경로는 별도 처리 필요
                "개성(퍼스널리티)": str(row.get('Personalities', '')).strip(),  # 새로 추가
                "속성명리스트": "",  # 속성 매핑 필요
                "속성_아이콘경로리스트": "",
                "무기명리스트": "",  # 무기 매핑 필요
                "무기_아이콘경로리스트": "",
                "출시일": str(row.get('Release Date', '')).strip(),
            })
        
        # CSV 저장
        csv_path = os.path.join(output_dir, "eden_roulette_data_with_personalities.csv")
        pd.DataFrame(result).to_csv(csv_path, index=False, encoding="utf-8-sig")
        log_queue_ref.put(f"CSV file generated: {csv_path}")
        
    except Exception as e:
        log_queue_ref.put(f"Error generating CSV: {e}")

# GUI 클래스 (기존과 유사하지만 Personalities 포함)
class PersonalityScraperApp:
    def __init__(self, root_window):
        self.root = root_window
        root_window.title("어나더에덴 캐릭터 스크레이퍼 (Personalities 포함)")
        root_window.geometry("900x700")

        load_config()
        self.output_dir_var = tk.StringVar(value=OUTPUT_BASE_DIR)

        style = ttk.Style()
        try:
            if os.name == 'nt': 
                style.theme_use('vista')
            else: 
                style.theme_use('clam') 
        except tk.TclError: 
            style.theme_use('clam')

        main_frame = ttk.Frame(root_window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 경로 설정
        path_frame = ttk.LabelFrame(main_frame, text="출력 설정", padding="10")
        path_frame.pack(fill=tk.X, pady=(0,10))
        
        ttk.Label(path_frame, text="저장 폴더:").grid(row=0, column=0, padx=(0,5), pady=5, sticky="w")
        self.output_dir_entry = ttk.Entry(path_frame, textvariable=self.output_dir_var, width=60)
        self.output_dir_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        self.btn_browse_dir = ttk.Button(path_frame, text="폴더 찾아보기", command=self.browse_output_directory)
        self.btn_browse_dir.grid(row=0, column=2, padx=(5,0), pady=5)
        path_frame.columnconfigure(1, weight=1)

        # 컨트롤
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0,10))

        self.btn_scrape = ttk.Button(
            control_frame, 
            text="🔍 캐릭터 데이터 수집 (Personalities 포함)", 
            command=self.start_scraping_thread
        )
        self.btn_scrape.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        
        self.btn_open_folder = ttk.Button(control_frame, text="📁 출력 폴더 열기", command=self.open_output_folder)
        self.btn_open_folder.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        # 로그
        log_label_frame = ttk.LabelFrame(main_frame, text="진행 상황 및 로그", padding="10")
        log_label_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(
            log_label_frame, wrap=tk.WORD, state=tk.DISABLED, height=20, 
            font=("Malgun Gothic", 9) if os.name == 'nt' else ("TkDefaultFont", 10)
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text.tag_configure("error", foreground="red")
        self.log_text.tag_configure("info", foreground="blue")
        self.log_text.tag_configure("success", foreground="green")

        self.progress_bar = ttk.Progressbar(log_label_frame, orient="horizontal", length=300, mode="determinate")
        self.progress_bar.pack(pady=10, fill=tk.X, padx=5)

        self.log_queue = queue.Queue()
        self.progress_queue = queue.Queue()
        self.root.after(100, self.process_queues)
        
        root_window.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        save_config()
        self.root.destroy()

    def browse_output_directory(self):
        global OUTPUT_BASE_DIR
        selected_dir = filedialog.askdirectory(initialdir=self.output_dir_var.get())
        if selected_dir:
            self.output_dir_var.set(selected_dir)
            OUTPUT_BASE_DIR = selected_dir
            self.log_message(f"출력 폴더가 다음으로 설정되었습니다: {selected_dir}", "info")
            save_config()

    def open_output_folder(self):
        folder_to_open = self.output_dir_var.get()
        if not os.path.isdir(folder_to_open):
            messagebox.showerror("오류", f"폴더를 찾을 수 없습니다:\n{folder_to_open}")
            return
        try:
            if os.name == 'nt': 
                os.startfile(folder_to_open)
            elif sys.platform == "darwin": 
                os.system(f'open "{folder_to_open}"')
            else: 
                os.system(f'xdg-open "{folder_to_open}"')
        except Exception as e:
            messagebox.showerror("폴더 열기 오류", f"폴더를 여는 중 오류 발생:\n{e}")

    def log_message(self, message, tag=None):
        self.log_text.config(state=tk.NORMAL)
        timestamp = time.strftime("%H:%M:%S", time.localtime())
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", tag)
        self.log_text.config(state=tk.DISABLED)
        self.log_text.see(tk.END) 

    def start_scraping_thread(self):
        current_output_dir = self.output_dir_var.get()
        if not os.path.isdir(current_output_dir):
            messagebox.showerror("오류", f"유효한 출력 폴더를 선택해주세요.")
            return

        self.btn_scrape.config(state=tk.DISABLED)
        self.progress_bar["value"] = 0
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete('1.0', tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.log_message("Personalities 포함 스크레이핑 시작...", "info")

        save_config()

        self.thread = threading.Thread(
            target=scraping_logic_with_personalities, 
            args=(self.log_queue, self.progress_queue, current_output_dir), 
            daemon=True
        )
        self.thread.start()

    def process_queues(self):
        try:
            while True: 
                message = self.log_queue.get_nowait()
                tag = "error" if "Error" in message or "오류" in message else "success" if "saved" in message or "완료" in message else "info"
                self.log_message(message, tag)
        except queue.Empty: 
            pass

        try:
            while True: 
                progress_data = self.progress_queue.get_nowait()
                if 'max' in progress_data: 
                    self.progress_bar["maximum"] = progress_data['max']
                if 'value' in progress_data: 
                    self.progress_bar["value"] = progress_data['value']
                if progress_data.get('done'):
                    self.btn_scrape.config(state=tk.NORMAL)
                    if self.progress_bar["value"] < self.progress_bar["maximum"]:
                        self.progress_bar["value"] = self.progress_bar["maximum"] 
                    
                    if progress_data.get('error'):
                        msg = progress_data.get('error_message', "작업 중 오류 발생.")
                        messagebox.showerror("오류", msg)
                    else:
                        messagebox.showinfo("완료", "Personalities 포함 데이터 수집이 완료되었습니다!")
        except queue.Empty: 
            pass
        
        self.root.after(100, self.process_queues)

if __name__ == "__main__":
    root = tk.Tk()
    app = PersonalityScraperApp(root)
    root.mainloop()