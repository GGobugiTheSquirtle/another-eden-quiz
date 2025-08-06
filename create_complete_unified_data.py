#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
완전한 통합 데이터 생성기
레거시 퍼스널리티 데이터를 활용하여 완전한 통합 CSV 생성
"""

import pandas as pd
import os
import re
import requests
import time
import mimetypes
import shutil
from pathlib import Path
from urllib.parse import urljoin, unquote, parse_qs, urlparse
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

# 프로젝트 루트 설정
PROJECT_ROOT = Path(__file__).parent.resolve()
CSV_DIR = PROJECT_ROOT / "04_data" / "csv"
IMAGE_DIR = PROJECT_ROOT / "04_data" / "images" / "character_art"
BASE_URL = "https://anothereden.wiki"
TARGET_URL = "https://anothereden.wiki/w/Characters"

# 무기/속성 번역 매핑 추가
WEAPON_TRANSLATION = {
    'Sword': '검',
    'Katana': '도',
    'Ax': '도끼',
    'Lance': '창',
    'Bow': '활',
    'Fists': '권갑',
    'Hammer': '망치',
    'Staff': '지팡이'
}

ELEMENT_TRANSLATION = {
    'Fire': '불',
    'Water': '물',
    'Wind': '바람',
    'Earth': '땅',
    'Thunder': '번개',
    'Shade': '그림자',
    'Crystal': '결정'
}

def translate_weapon_element(text):
    """무기와 속성을 한국어로 번역"""
    if not text:
        return text
    
    # 무기 번역
    for eng, kor in WEAPON_TRANSLATION.items():
        text = text.replace(eng, kor)
    
    # 속성 번역
    for eng, kor in ELEMENT_TRANSLATION.items():
        text = text.replace(eng, kor)
    
    return text

def download_image(image_url, subfolder=""):
    """레거시 방식으로 이미지 다운로드"""
    if not image_url: 
        return None
    
    full_image_url = urljoin(BASE_URL, image_url)
    try:
        parsed_url = urlparse(full_image_url)
        query_params = parse_qs(parsed_url.query)
        image_name_from_f = query_params.get('f', [None])[0]
        
        if image_name_from_f:
            image_name = os.path.basename(unquote(image_name_from_f))
        else:
            image_name = os.path.basename(unquote(parsed_url.path.split('?')[0]))

        if not image_name or image_name.lower() in ["thumb.php", "index.php"]:
            temp_name = unquote(full_image_url.split('/')[-1].split('?')[0])
            image_name = (temp_name[:50] + ".png") if temp_name else "unknown_image.png"

        base_name, ext = os.path.splitext(image_name)
        if not ext or len(ext) > 5:
            try:
                head_resp = requests.head(full_image_url, timeout=3, allow_redirects=True)
                head_resp.raise_for_status()
                content_type = head_resp.headers.get('Content-Type')
                if content_type:
                    guessed_ext = mimetypes.guess_extension(content_type.split(';')[0])
                    image_name = base_name + (guessed_ext if guessed_ext and guessed_ext != '.jpe' else ('.jpg' if guessed_ext == '.jpe' else '.png'))
                else: 
                    image_name = base_name + ".png" 
            except: 
                image_name = base_name + ".png"
        
        image_name = re.sub(r'[<>:"/\\|?*]', '_', image_name)
        image_name = image_name[:200]

        save_path_dir = os.path.join(IMAGE_DIR, subfolder)
        save_path = os.path.join(save_path_dir, image_name)
        
        if os.path.exists(save_path) and os.path.getsize(save_path) > 0:
            return save_path

        img_response = requests.get(full_image_url, stream=True, timeout=10)
        img_response.raise_for_status()
        with open(save_path, 'wb') as f:
            for chunk in img_response.iter_content(8192):
                f.write(chunk)
        time.sleep(0.05) 
        return save_path
    except Exception as e:
        print(f"Download Error for {full_image_url.split('/')[-1][:30]}...: {type(e).__name__}")
    return None

def organize_images(character_data, output_dir):
    """레거시 방식으로 이미지 정리 (출시일순/가나다순)"""
    print("📁 이미지 정리 시작...")
    
    # 정리된 이미지 폴더 생성
    organized_base_dir = os.path.join(output_dir, "캐릭터 정리")
    by_date_dir = os.path.join(organized_base_dir, "1. 출시일 순")
    by_name_dir = os.path.join(organized_base_dir, "2. 가나다 순")
    os.makedirs(by_date_dir, exist_ok=True)
    os.makedirs(by_name_dir, exist_ok=True)

    # 출시일 순으로 정렬 (날짜가 없는 경우 맨 뒤로 보내기 위해 '9999/99/99' 사용)
    sorted_by_date = sorted(character_data, key=lambda x: x.get('release_date', '9999/99/99'))
    
    # 가나다 순으로 정렬
    sorted_by_name = sorted(character_data, key=lambda x: x.get('name', ''))

    def copy_and_rename_image(char_data, index, target_dir):
        """이미지 복사 및 이름 변경"""
        if char_data["icon_path"] and os.path.exists(char_data["icon_path"]):
            try:
                # 희귀도 변환 (5★ SA → 5성성각)
                rarity_transformed = char_data['rarity'].strip().replace(" ", "").replace("★", "성").replace("SA", "성각") or "정보없음"
                
                # 출시일 변환 (YYYY/MM/DD → YYYY-MM-DD)
                date_transformed = char_data['release_date'].replace('/', '-') if char_data['release_date'] else "날짜없음"
                
                # 기본 파일명 생성
                base_filename = f"{char_data['name']}_{rarity_transformed}_{date_transformed}"
                sanitized_filename_base = re.sub(r'[<>:"/\\|?*]', '_', base_filename)
                
                # 확장자 처리
                _, extension = os.path.splitext(char_data["icon_path"])
                if not extension: 
                    extension = '.png'
                
                # 번호(넘버링) 추가
                final_organized_filename = f"{index+1:03d}_{sanitized_filename_base}{extension}"
                
                destination_path = os.path.join(target_dir, final_organized_filename)
                shutil.copy2(char_data["icon_path"], destination_path)
                
                if index % 50 == 0:
                    print(f"   이미지 정리 진행률: {index+1}/{len(character_data)}")
                    
            except Exception as e:
                print(f"Failed to copy organized image for '{char_data['name']}': {e}")

    # 정렬된 목록을 기반으로 파일 복사 및 이름 변경 실행
    print("   출시일순 정리 중...")
    for idx, data in enumerate(sorted_by_date):
        copy_and_rename_image(data, idx, by_date_dir)
    
    print("   가나다순 정리 중...")
    for idx, data in enumerate(sorted_by_name):
        copy_and_rename_image(data, idx, by_name_dir)
        
    print(f"✅ 이미지 정리 완료: {organized_base_dir}")
    return organized_base_dir

def make_roulette_csv_from_excel(src_excel="another_eden_characters_detailed.xlsx"):
    """레거시 방식으로 Excel에서 룰렛용 CSV 생성"""
    print(f"🎲 룰렛용 CSV 생성 중: {src_excel}")
    
    try:
        df = pd.read_excel(src_excel, engine='openpyxl')
    except FileNotFoundError:
        print(f"❌ 입력 파일 없음: {src_excel}")
        return None

    # 이름 매칭용 CSV 파일을 읽고, 변환 규칙(딕셔너리)을 생성합니다.
    try:
        name_map_df = pd.read_csv("Matching_names.csv")
        name_map = dict(zip(name_map_df["캐릭터명 (입력)"], name_map_df["캐릭터명 (매칭)"]))
    except FileNotFoundError:
        print("⚠️ 이름 매칭 파일(Matching_names.csv)을 찾을 수 없습니다. 캐릭터명이 영문으로 표시됩니다.")
        name_map = {}
    except Exception as e:
        print(f"❌ 이름 매칭 파일(Matching_names.csv) 로드 중 오류 발생: {e}")
        return None

    # 컬럼 찾기 함수
    def pick_col(df, candidates, exact_match=False):
        for candidate in candidates:
            if exact_match:
                if candidate in df.columns:
                    return candidate
            else:
                for col in df.columns:
                    if candidate.lower() in col.lower():
                        return col
        return None

    # HTML 태그 정리 함수
    def clean_html_tags(text):
        if pd.isna(text):
            return ""
        return re.sub(r'<[^>]+>', '', str(text)).strip()

    # 경로 정리 함수
    def clean_path_list(lst):
        return [str(p) for p in lst if p and str(p).strip()]

    # 이미지 찾기 함수
    def find_image(filename, subdir, character_name=None):
        if not filename or pd.isna(filename):
            return ""
        
        # 상대 경로로 변환
        base_path = f"04_data/images/character_art/{subdir}/{filename}"
        if os.path.exists(base_path):
            return base_path
        return ""

    # Buddy 장비 필터링 (레거시 방식)
    BUDDY_PATTERN = re.compile(r'Buddy equipment\.png', re.IGNORECASE)
    
    buddy_check_cols = [pick_col(df, [f"Elem/Equip {i} Alt"]) for i in range(1, 6)]
    buddy_check_cols = [c for c in buddy_check_cols if c]

    def is_buddy_row(row):
        for col_name in buddy_check_cols:
            val = str(row.get(col_name, ""))
            if BUDDY_PATTERN.search(val):
                return True
        return False
    
    if buddy_check_cols:
        df_clean = df[~df.apply(is_buddy_row, axis=1)].reset_index(drop=True)
        print(f"✅ Buddy 장비 필터링 완료: {len(df)} → {len(df_clean)}개")
    else:
        print("⚠️ Buddy 장비 필터링을 위한 'Elem/Equip * Alt' 컬럼을 찾지 못했습니다. 전체 데이터가 사용됩니다.")
        df_clean = df.copy()

    # 컬럼 찾기
    col_icon_filename = pick_col(df_clean, ["Icon Filename", "Icon", "아이콘 파일명"], exact_match=True)
    col_name = pick_col(df_clean, ["Name", "이름"], exact_match=True)
    col_rarity = pick_col(df_clean, ["Rarity", "희귀도"], exact_match=True)
    col_release = pick_col(df_clean, ["Release Date", "출시일"], exact_match=True)

    if not all([col_icon_filename, col_name, col_rarity, col_release]):
        missing_cols = [c_name for c_name, c_val in zip(
            ["아이콘 파일명", "이름", "희귀도", "출시일"],
            [col_icon_filename, col_name, col_rarity, col_release]) if not c_val]
        print(f"❌ 필수 컬럼 누락: {', '.join(missing_cols)}. Excel 파일을 확인해주세요.")
        return None

    # 한글 이름 변환 함수
    def find_best_match_and_translate(english_name, name_map):
        if not english_name or pd.isna(english_name):
            return english_name
        
        english_name = str(english_name).strip()
        
        # 정확한 매칭
        if english_name in name_map:
            return name_map[english_name]
        
        # 부분 매칭 (AS, ES 등 포함)
        for eng_key, kor_value in name_map.items():
            if english_name in eng_key or eng_key in english_name:
                return kor_value
        
        return english_name

    # 속성/무기/방어구 매핑 (레거시 방식)
    skillType_map = {
        "Fire.png": "Fire",
        "Water.png": "Water", 
        "Wind.png": "Wind",
        "Earth.png": "Earth",
        "Crystal.png": "Crystal",
        "Shade.png": "Shade",
        "Thunder.png": "Thunder"
    }
    
    weapon_map = {
        "Sword.png": "Sword",
        "Katana.png": "Katana",
        "Ax.png": "Ax",
        "Lance.png": "Lance", 
        "Bow.png": "Bow",
        "Fists.png": "Fists",
        "Hammer.png": "Hammer",
        "Staff.png": "Staff"
    }
    
    armor_map = {
        "Light_Armor.png": "Light Armor",
        "Heavy_Armor.png": "Heavy Armor"
    }

    result = []
    for idx, row in df_clean.iterrows():
        # 영문 이름을 가져옵니다.
        name = str(row.get(col_name, "")).strip()
        
        # 한글 이름 변환
        korean_name = find_best_match_and_translate(name, name_map)
        
        icon_file = str(row.get(col_icon_filename, "")).strip().replace(" ", "_")
        rarity_raw = str(row.get(col_rarity, "")).strip()
        rarity = clean_html_tags(rarity_raw)
        
        release_raw = str(row.get(col_release, "")).strip()
        release = clean_html_tags(release_raw)

        icon_path = find_image(icon_file, "icons", name)
        if not icon_path and icon_file:
            print(f"⚠️ [캐릭터 아이콘 없음] 파일: {icon_file} (캐릭터: {name})")

        attr_names, attr_paths = [], []
        weapon_names, weapon_paths = [], []
        armor_names, armor_paths = [], []

        # 속성/장비 아이콘 처리 개선 - 한글 컬럼명 처리
        elem_equip_cols = []
        
        # 한글 컬럼명으로 속성/장비 정보 추출
        for i in range(1, 6):
            info_col = f"속성/장비 정보 {i}"
            if info_col in df_clean.columns:
                val = str(row.get(info_col, "")).strip()
                if val and val != 'nan':
                    elem_equip_cols.append(val)
        
        # 영어 컬럼명으로도 시도
        if not elem_equip_cols:
            for i in range(1, 6):
                ecol = pick_col(df_clean, [f"Elem/Equip {i} Alt", f"Elem/Equip{i}Alt"], exact_match=False)
                if ecol:
                    val = str(row.get(ecol, "")).strip()
                    if val and val != 'nan':
                        elem_equip_cols.append(val)
        
        # 실제 파일명 기반 매핑
        for val in elem_equip_cols:
            if not val or pd.isna(val) or val == 'nan':
                continue
                
            val_norm = val.strip().replace(" ", "_")
            
            # 파일명에서 확장자 제거
            if val_norm.endswith('.png'):
                val_norm = val_norm[:-4]
            
            # 속성 매핑
            if val_norm in skillType_map:
                attr_names.append(skillType_map[val_norm])
                attr_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            # 무기 매핑
            elif val_norm in weapon_map:
                weapon_names.append(weapon_map[val_norm])
                weapon_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            # 방어구 매핑
            elif val_norm in armor_map:
                armor_names.append(armor_map[val_norm])
                armor_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            # 추가 매핑 (실제 파일명 기반)
            elif "fire" in val_norm.lower():
                attr_names.append("Fire")
                attr_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            elif "water" in val_norm.lower():
                attr_names.append("Water")
                attr_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            elif "wind" in val_norm.lower():
                attr_names.append("Wind")
                attr_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            elif "earth" in val_norm.lower():
                attr_names.append("Earth")
                attr_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            elif "sword" in val_norm.lower():
                weapon_names.append("Sword")
                weapon_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            elif "katana" in val_norm.lower():
                weapon_names.append("Katana")
                weapon_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            elif "bow" in val_norm.lower():
                weapon_names.append("Bow")
                weapon_paths.append(find_image(val_norm + ".png", "elements_equipment", name))
            elif "staff" in val_norm.lower():
                weapon_names.append("Staff")
                weapon_paths.append(find_image(val_norm + ".png", "elements_equipment", name))

        result.append({
            "캐릭터명": korean_name,
            "희귀도": rarity,
            "캐릭터아이콘경로": icon_path or "",
            "속성명리스트": ",".join(attr_names),
            "속성_아이콘경로리스트": ",".join(clean_path_list(attr_paths)),
            "무기명리스트": ",".join(weapon_names),
            "무기_아이콘경로리스트": ",".join(clean_path_list(weapon_paths)),
            "방어구명리스트": ",".join(armor_names),
            "방어구_아이콘경로리스트": ",".join(clean_path_list(armor_paths)),
            "출시일": release,
        })
        
    re_df = pd.DataFrame(result)
    out_csv = CSV_DIR / "eden_roulette_data_from_excel.csv"
    re_df.to_csv(out_csv, index=False, encoding="utf-8-sig")
    print(f"✅ 룰렛용 CSV 생성 완료: {out_csv}")
    return str(out_csv)

def scrape_character_table():
    """레거시 방식으로 캐릭터 테이블에서 데이터 스크래핑"""
    print("📡 캐릭터 테이블 스크래핑 시작...")
    
    headers_ua = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
    try:
        response = requests.get(TARGET_URL, headers=headers_ua, timeout=15)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching page: {e}")
        return []

    soup = BeautifulSoup(response.content, 'html.parser')
    
    char_table = soup.find('table', class_='chara-table') or soup.find('table', class_='wikitable')
    if not char_table:
        print("Character table not found. Please verify table class name.")
        return []

    print("Character table found. Parsing rows...")
    
    character_data = []
    rows = char_table.find_all('tr')
    total_rows = len(rows) - 1
    
    print(f"Found {total_rows} potential character rows.")
    
    for i, row in enumerate(rows[1:], 1):
        if i % 10 == 0:
            print(f"   진행률: {i}/{total_rows}")
        
        cells = row.find_all('td')
        if len(cells) < 4: 
            continue
        
        try:
            # 아이콘 셀 (cells[0])
            icon_cell = cells[0]
            icon_img_tag = icon_cell.find('img')
            icon_src = icon_img_tag['src'] if icon_img_tag and icon_img_tag.get('src') else None
            icon_local_path = download_image(icon_src, "icons") if icon_src else None
            
            icon_filename = ""
            if icon_src:
                parsed_url = urlparse(icon_src)
                query_params = parse_qs(parsed_url.query)
                image_name_from_f = query_params.get('f', [None])[0]
                if image_name_from_f: 
                    icon_filename = os.path.basename(unquote(image_name_from_f)).replace(' ', '_')
                else: 
                    icon_filename = os.path.basename(unquote(parsed_url.path.split('?')[0])).replace(' ', '_')

            # 이름/희귀도 셀 (cells[1])
            name_rarity_cell = cells[1]
            name_tag = name_rarity_cell.find('a')
            
            original_name = name_tag.text.strip() if name_tag else ""
            rarity = ""
            lines_in_cell = [line.strip() for line in name_rarity_cell.get_text(separator='\n').splitlines() if line.strip()]
            if not original_name and lines_in_cell: 
                original_name = lines_in_cell[0]
            for line_text in reversed(lines_in_cell):
                if "★" in line_text:
                    rarity = line_text
                    break
            if not rarity:
                rarity_match = re.search(r'\d(?:~\d)?★(?:\s*\S+)?', name_rarity_cell.get_text(separator=" ").strip())
                if rarity_match: 
                    rarity = rarity_match.group(0).strip()
            
            # 속성/장비 셀 (cells[2])
            element_equipment_cell = cells[2]
            ee_icon_tags = element_equipment_cell.find_all('img')
            element_equipment_icon_paths = []
            element_equipment_icon_alts = []
            for img_tag in ee_icon_tags:
                ee_src = img_tag.get('src')
                ee_alt = img_tag.get('alt', "") 
                if ee_src:
                    local_path = download_image(ee_src, "elements_equipment")
                    if local_path:
                        element_equipment_icon_paths.append(str(Path(local_path).relative_to(PROJECT_ROOT)))
                        element_equipment_icon_alts.append(ee_alt)

            # 출시일 셀 (cells[3])
            release_date = cells[3].text.strip() if len(cells) > 3 else ""

            if original_name or icon_local_path:
                character_data.append({
                    "icon_path": str(Path(icon_local_path).relative_to(PROJECT_ROOT)) if icon_local_path else "",
                    "icon_filename": icon_filename,
                    "name": original_name,
                    "rarity": rarity,
                    "element_equipment_paths": element_equipment_icon_paths,
                    "element_equipment_alts": element_equipment_icon_alts,
                    "release_date": release_date
                })

        except Exception as e:
            print(f"Error parsing row {i}: {e}")
            continue
    
    print(f"✅ 테이블 스크래핑 완료: {len(character_data)}개 캐릭터")
    return character_data

def load_personality_data():
    """퍼스널리티 데이터 로드"""
    personality_file = Path("character_personalities.csv")
    if not personality_file.exists():
        print("❌ character_personalities.csv 파일이 없습니다")
        return {}
    
    try:
        df = pd.read_csv(personality_file)
        personality_data = {}
        
        for _, row in df.iterrows():
            eng_name = row['English_Name']
            personalities = row['Personalities_List'].split('|')
            personality_data[eng_name] = personalities
        
        print(f"✅ 퍼스널리티 데이터 로드: {len(personality_data)}명")
        return personality_data
    except Exception as e:
        print(f"❌ 퍼스널리티 데이터 로드 실패: {e}")
        return {}

def load_name_mapping():
    """한글 이름 매핑 로드"""
    mapping_file = CSV_DIR / "Matching_names.csv"
    if not mapping_file.exists():
        print("❌ Matching_names.csv 파일이 없습니다")
        return {}
    
    try:
        df = pd.read_csv(mapping_file, encoding='utf-8-sig')
        name_mapping = {}
        
        for _, row in df.iterrows():
            if len(row) >= 2:
                eng_name = str(row.iloc[0]).strip()
                kor_name = str(row.iloc[1]).strip()
                if eng_name and kor_name and kor_name != 'nan':
                    name_mapping[eng_name] = kor_name
        
        print(f"✅ 한글 매핑 로드: {len(name_mapping)}개")
        return name_mapping
    except Exception as e:
        print(f"❌ 한글 매핑 로드 실패: {e}")
        return {}

def extract_elements_and_weapons(personalities):
    """퍼스널리티에서 속성과 무기 추출"""
    elements = []
    weapons = []
    
    # 속성 정의
    element_keywords = ['Fire', 'Water', 'Wind', 'Earth', 'Crystal', 'Shade', 'Thunder']
    # 무기 정의
    weapon_keywords = ['Sword', 'Katana', 'Ax', 'Lance', 'Bow', 'Fists', 'Hammer', 'Staff']
    
    for personality in personalities:
        if personality in element_keywords:
            elements.append(personality)
        elif personality in weapon_keywords:
            weapons.append(personality)
    
    return elements, weapons

def create_excel_file(character_data, output_path):
    """레거시 방식으로 Excel 파일 생성"""
    print(f"📊 Excel 파일 생성 중: {output_path}")
    
    # 헤더 매핑
    HEADER_MAP = {
        "Icon": "아이콘", "Icon Filename": "아이콘 파일명", "Name": "이름", "Rarity": "희귀도",
        "Elem/Equip Icon": "속성/장비", "Elem/Equip Alt": "속성/장비 정보", "Release Date": "출시일"
    }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Characters"
    
    # 최대 속성/장비 아이콘 수 계산
    max_ee_icons = max((len(d["element_equipment_paths"]) for d in character_data), default=0)
    
    # 헤더 생성
    headers_excel = [HEADER_MAP["Icon"], HEADER_MAP["Icon Filename"], HEADER_MAP["Name"], HEADER_MAP["Rarity"]]
    for i in range(max_ee_icons):
        headers_excel.extend([f"{HEADER_MAP['Elem/Equip Icon']} {i+1}", f"{HEADER_MAP['Elem/Equip Alt']} {i+1}"])
    headers_excel.append(HEADER_MAP["Release Date"])
    ws.append(headers_excel)

    # 컬럼 너비 설정
    ws.column_dimensions[get_column_letter(1)].width = 12 
    ws.column_dimensions[get_column_letter(2)].width = 35 
    ws.column_dimensions[get_column_letter(3)].width = 25 
    ws.column_dimensions[get_column_letter(4)].width = 15
    col_offset = 5
    for i in range(max_ee_icons):
        ws.column_dimensions[get_column_letter(col_offset + i*2)].width = 12 
        ws.column_dimensions[get_column_letter(col_offset + i*2 + 1)].width = 25 
    ws.column_dimensions[get_column_letter(col_offset + max_ee_icons*2)].width = 15
    
    # 데이터 행 추가
    for row_idx_excel, char_data in enumerate(character_data, start=2):
        ws.row_dimensions[row_idx_excel].height = 60 
        current_col = 1 
        
        # 아이콘 이미지 삽입
        if char_data["icon_path"] and os.path.exists(char_data["icon_path"]):
            try:
                img = OpenpyxlImage(char_data["icon_path"])
                img.height = 75 
                img.width = 75  
                ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
            except Exception as e:
                print(f"Excel Write Error (Icon) for {char_data.get('name', 'N/A')}: {e}")
                ws.cell(row=row_idx_excel, column=current_col, value="ImgErr")
        current_col += 1
        
        # 아이콘 파일명
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("icon_filename", ""))
        current_col += 1
        
        # 이름
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("name", "N/A"))
        current_col += 1
        
        # 희귀도
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("rarity", "N/A"))
        current_col += 1

        # 속성/장비 아이콘들
        for i in range(max_ee_icons):
            if i < len(char_data["element_equipment_paths"]):
                icon_path = char_data["element_equipment_paths"][i]
                icon_alt = char_data["element_equipment_alts"][i]
                if icon_path and os.path.exists(icon_path):
                    try:
                        img = OpenpyxlImage(icon_path)
                        img.height = 30 
                        img.width = 30
                        ws.add_image(img, f'{get_column_letter(current_col)}{row_idx_excel}')
                    except Exception as e:
                        print(f"Excel Write Error (E/E Icon) for {char_data.get('name', 'N/A')}: {e}")
                        ws.cell(row=row_idx_excel, column=current_col, value="ImgErr")
                ws.cell(row=row_idx_excel, column=current_col + 1, value=icon_alt)
            else: 
                ws.cell(row=row_idx_excel, column=current_col, value="")
                ws.cell(row=row_idx_excel, column=current_col + 1, value="")
            current_col += 2 
        
        # 출시일
        ws.cell(row=row_idx_excel, column=current_col, value=char_data.get("release_date", "N/A"))
        
        if (row_idx_excel-1) % 50 == 0:
            print(f"Excel: Wrote data for '{char_data.get('name')}' (Processed {row_idx_excel-1} characters)")

    try:
        wb.save(output_path)
        print(f"✅ Excel 파일 생성 완료: {output_path}")
        return True
    except Exception as e:
        print(f"❌ Excel 파일 저장 실패: {e}")
        return False

def create_character_data():
    """캐릭터 데이터 생성"""
    print("🎮 완전한 통합 데이터 생성 시작")
    print("=" * 60)
    
    # 테이블에서 데이터 스크래핑
    table_data = scrape_character_table()
    
    # 데이터 로드
    personality_data = load_personality_data()
    name_mapping = load_name_mapping()
    
    if not table_data:
        print("❌ 테이블 데이터가 없어 생성할 수 없습니다")
        return
    
    # 통합 데이터 생성
    unified_data = []
    
    print("🔄 데이터 통합 시작...")
    for char_data in table_data:
        eng_name = char_data['name']
        
        # 한글 이름 변환 (매핑이 있으면 사용, 없으면 영어 이름 그대로)
        korean_name = name_mapping.get(eng_name, eng_name)
        
        # 퍼스널리티 데이터 가져오기
        personalities = personality_data.get(eng_name, [])
        
        # 속성과 무기 추출
        elements, weapons = extract_elements_and_weapons(personalities)
        
        # 한국어로 번역
        elements_korean = [translate_weapon_element(elem) for elem in elements]
        weapons_korean = [translate_weapon_element(weapon) for weapon in weapons]
        
        # 희귀도 (테이블에서 가져온 값 사용)
        rarity = char_data['rarity']
        
        # 기본 속성/무기 아이콘 경로 (퍼스널리티 기반)
        element_icons = []
        weapon_icons = []
        
        for element in elements:
            element_icon = f"04_data/images/character_art/elements_equipment/{element.lower()}_icon.png"
            element_icons.append(element_icon)
        
        for weapon in weapons:
            weapon_icon = f"04_data/images/character_art/elements_equipment/{weapon.lower()}_icon.png"
            weapon_icons.append(weapon_icon)
        
        # 통합 데이터 행 생성
        unified_row = {
            '캐릭터명': korean_name,
            'English_Name': eng_name,
            '캐릭터아이콘경로': char_data['icon_path'],
            '희귀도': rarity,
            '속성명리스트': ', '.join(elements_korean),
            '무기명리스트': ', '.join(weapons_korean),
            '퍼스널리티리스트': ', '.join(personalities),
            '속성_아이콘경로리스트': '|'.join(element_icons),
            '무기_아이콘경로리스트': '|'.join(weapon_icons),
            '방어구_아이콘경로리스트': '|'.join(char_data['element_equipment_paths']),
            '속성장비_정보': '|'.join(char_data['element_equipment_alts']),
            '출시일': char_data['release_date']
        }
        
        unified_data.append(unified_row)
    
    print("✅ 데이터 생성 완료!")
    
    # CSV 생성 - 앱에서 사용하는 파일명으로 생성
    df = pd.DataFrame(unified_data)
    
    # 1. 통합 데이터 (eden_unified_data.csv)
    unified_csv_path = CSV_DIR / "eden_unified_data.csv"
    df.to_csv(unified_csv_path, index=False, encoding='utf-8-sig')
    print(f"✅ 통합 CSV 생성 완료: {unified_csv_path}")
    
    # 2. 퀴즈용 데이터 (eden_quiz_data.csv) - 앱에서 우선 사용
    quiz_csv_path = CSV_DIR / "eden_quiz_data.csv"
    df.to_csv(quiz_csv_path, index=False, encoding='utf-8-sig')
    print(f"✅ 퀴즈용 CSV 생성 완료: {quiz_csv_path}")
    
    # 3. 룰렛용 데이터 (eden_roulette_data.csv) - 앱에서 대안으로 사용
    roulette_csv_path = CSV_DIR / "eden_roulette_data.csv"
    df.to_csv(roulette_csv_path, index=False, encoding='utf-8-sig')
    print(f"✅ 룰렛용 CSV 생성 완료: {roulette_csv_path}")
    
    # 4. Excel 파일 생성 (레거시 방식)
    excel_path = CSV_DIR / "another_eden_characters_detailed.xlsx"
    create_excel_file(table_data, excel_path)
    
    # 5. Excel에서 룰렛용 CSV 생성 (레거시 방식)
    roulette_from_excel_path = make_roulette_csv_from_excel(str(excel_path))
    
    # 6. 이미지 정리 (레거시 방식)
    organized_dir = organize_images(table_data, str(CSV_DIR))
    
    print(f"📊 처리된 캐릭터: {len(unified_data)}개")
    
    # 통계 출력
    print("\n📈 데이터 통계:")
    print(f"   🎭 총 캐릭터: {len(unified_data)}명")
    print(f"   🔥 속성 포함 캐릭터: {len([c for c in unified_data if c['속성명리스트']])}명")
    print(f"   ⚔️ 무기 포함 캐릭터: {len([c for c in unified_data if c['무기명리스트']])}명")
    print(f"   💎 5성 캐릭터: {len([c for c in unified_data if '5★' in c['희귀도']])}명")
    print(f"   🌟 4성 캐릭터: {len([c for c in unified_data if '4★' in c['희귀도']])}명")
    print(f"   ⭐ 3성 캐릭터: {len([c for c in unified_data if '3★' in c['희귀도']])}명")
    
    # 한글 매핑 통계
    mapped_count = len([c for c in unified_data if c['캐릭터명'] != c['English_Name']])
    print(f"   🇰🇷 한글 매핑: {mapped_count}명")
    
    # 이미지 다운로드 통계
    downloaded_icons = len([c for c in unified_data if c['방어구_아이콘경로리스트']])
    print(f"   🖼️ 다운로드된 아이콘: {downloaded_icons}개")
    
    # 샘플 데이터 출력
    print("\n📋 샘플 데이터 (처음 5개):")
    for i, char in enumerate(unified_data[:5], 1):
        print(f"   {i}. {char['캐릭터명']} ({char['English_Name']})")
        print(f"      희귀도: {char['희귀도']}")
        print(f"      속성: {char['속성명리스트']}")
        print(f"      무기: {char['무기명리스트']}")
        print(f"      퍼스널리티: {char['퍼스널리티리스트'][:50]}...")
        print()
    
    return [unified_csv_path, quiz_csv_path, roulette_csv_path, excel_path, roulette_from_excel_path, organized_dir]

def main():
    """메인 실행 함수"""
    # 디렉토리 생성
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    (IMAGE_DIR / "icons").mkdir(exist_ok=True)
    (IMAGE_DIR / "elements_equipment").mkdir(exist_ok=True)
    
    # 통합 데이터 생성
    csv_paths = create_character_data()
    
    if csv_paths:
        print("\n🎉 완전한 통합 데이터 생성 완료!")
        print("=" * 60)
        print("💾 생성된 파일들:")
        for path in csv_paths:
            print(f"   📄 {path}")
        print("\n🚀 이제 퀴즈 앱과 룰렛 앱에서 사용할 수 있습니다!")
        return True
    else:
        print("\n❌ 통합 데이터 생성 실패")
        return False

if __name__ == "__main__":
    success = main()
    exit(0 if success else 1) 